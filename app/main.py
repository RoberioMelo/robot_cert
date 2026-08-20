from __future__ import annotations

import logging
import csv
import html
import io
import json
import os
import re
import threading
import time
import unicodedata
from collections import defaultdict
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from difflib import SequenceMatcher
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote

from fastapi import Depends, FastAPI, File, Header, HTTPException, Query, Request, UploadFile, BackgroundTasks
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, Field

from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer

from app import atividade, auth, config, permissoes, senha_reset
from app.historico_agg_cache import get_or_build as _historico_cache_get_or_build
from app.cert_scanner import CertInfo, CertStatus, cert_to_public_dict, move_to_expired, scan_folder
from app.command_queue import COMMANDS, enqueue, list_pending, pop_next_for_agent
from app.config import ROOT
from app import smtp_service
from app.smtp_service import encrypt_password, validate_smtp_config
from app.alert_state import trigger_all_alerts, job_ja_executado_recentemente
from app.notification_service import build_notifications_payload
from app.settings_state import (
    PortalSettings,
    get_latest_snapshot,
    load_colaborador_selecao,
    load_settings,
    save_colaborador_selecao,
    save_settings,
    save_snapshot,
    supabase_configured,
    upsert_cert_history,
)

security = HTTPBearer(auto_error=False)

# Identidade atribuída quando API_KEY não está configurada (ambiente aberto).
# Rotas sensíveis devem recusá-la — é o oposto de "autenticado".
ANONYMOUS_IDENTITY_EMAIL = "anonymous@local"

# Mensagem única para "este token não vale mais". Não distingue conta excluída
# de conta desativada de propósito: quem está do outro lado já perdeu o acesso,
# e a diferença só serviria para dizer a um ex-usuário em que estado a conta
# dele ficou.
SESSAO_ENCERRADA = "Sessão encerrada. Entre novamente."

# Rotas que continuam funcionando com senha provisória. Lista fechada, e
# curta de propósito: tudo o mais é recusado. Fosse uma lista de bloqueio em
# vez de liberação, cada rota nova nasceria acessível por omissão — e o
# esquecimento não daria sintoma nenhum.
ROTAS_COM_SENHA_PROVISORIA = frozenset({"/api/senha/trocar"})

ERRO_SENHA_PROVISORIA = (
    "Sua senha foi definida por outra pessoa. Escolha uma senha própria para "
    "continuar."
)


class ContaIndisponivel(RuntimeError):
    """O diretório de usuários existe, mas não respondeu."""


class ContaInvalida(RuntimeError):
    """A conta que o token nomeia não existe mais no diretório."""


def _conta_da_sessao(email: str) -> Optional[dict]:
    """
    Relê a conta a cada requisição, para o token não congelar a permissão.

    Devolve `None` quando **não há** diretório de usuários configurado — dev e
    testes sem Supabase. Aí não existe conta contra a qual conferir, e o token é
    a única informação disponível. Isso não abre brecha em produção: sem
    Supabase o `/api/login` responde 503 e ninguém chega a ter um token para
    apresentar.

    Levanta `ContaInvalida` quando a conta sumiu — excluída, ou com o e-mail
    alterado, porque aí o `sub` do token deixa de casar com qualquer linha.
    Levanta `ContaIndisponivel` quando a leitura falha. Barrar no segundo caso é
    deliberado: "não consegui verificar" não é "está tudo certo". É a mesma
    escolha já feita em `CustodiaIndisponivel`, e pela mesma razão — a variante
    permissiva não daria sintoma nenhum.
    """
    from app.settings_state import _supabase

    sb = _supabase()
    if not sb:
        return None
    try:
        r = (
            sb.table("users")
            # `senha_alterada_em` PRECISA estar aqui: é o que
            # `_senha_trocada_depois_do_token` compara. Omiti-la não daria erro
            # — a chave chegaria ausente, seria lida como "nunca trocou", e a
            # invalidação de sessão simplesmente nunca dispararia. Fixado por
            # `test_sessao_le_a_coluna_da_troca_de_senha`, porque o fake dos
            # testes devolve a linha inteira e não pega isto sozinho.
            .select("id, email, role, ativo, senha_alterada_em, deve_trocar_senha")
            .eq("email", email)
            .limit(1)
            .execute()
        )
    except Exception as e:
        logger.warning("Não foi possível verificar a conta da sessão: %s", e)
        raise ContaIndisponivel(str(e)) from e
    if not r.data:
        raise ContaInvalida(email)
    return r.data[0]


def _senha_trocada_depois_do_token(
    conta: dict, token_data: auth.TokenData
) -> bool:
    """
    A senha mudou depois de este token ter sido emitido?

    Devolve False quando falta informação — coluna nula (conta que nunca trocou
    a senha), token sem instante de emissão, ou data ilegível. É o único ponto
    fail-OPEN do `_sessao_do_token`, e é deliberado: um erro de leitura aqui
    deslogaria o portal inteiro de uma vez, e a ausência de dado significa
    literalmente "não houve troca a invalidar".

    A margem de 5s absorve relógios ligeiramente fora de sincronia entre o
    processo que emitiu o token e o que gravou a troca. Sem ela, uma diferença
    de milissegundos poderia derrubar a sessão de quem acabou de entrar.
    """
    marca = conta.get("senha_alterada_em")
    if not marca or not token_data.emitido_em:
        return False
    try:
        trocada = _parse_iso_utc(str(marca))
    except Exception:  # noqa: BLE001
        logger.warning("senha_alterada_em ilegível para %s", conta.get("email"))
        return False
    if trocada is None:
        return False
    return token_data.emitido_em < (trocada - timedelta(seconds=5))


def _sessao_do_token(token_data: auth.TokenData) -> auth.TokenData:
    """
    O papel que vale é o do banco, não o que veio dentro do token.

    Sem isto, `require_admin` decidia com `token.role` — gravado no login e
    válido por 24h (`auth.ACCESS_TOKEN_EXPIRE_MINUTES`). Duas consequências, e a
    segunda é a pior: desativar alguém não derrubava a sessão aberta dele, e
    **rebaixar um administrador não tirava o poder de administrador**; os dois
    efeitos só chegavam quando o token expirasse.

    Era tolerável enquanto desativar era operação rara. Com a hierarquia
    gestor/operador de 15/08, desativar passou a ser a forma principal de
    revogar acesso — e revogação que leva um dia para valer não é revogação.

    Custa uma leitura em `users` por requisição autenticada. É o preço honesto:
    a alternativa por `token_version` faria a mesma leitura, só que precisando
    também de migration.
    """
    try:
        conta = _conta_da_sessao(token_data.email)
    except ContaInvalida:
        raise HTTPException(
            status_code=401,
            detail=SESSAO_ENCERRADA,
            headers={"WWW-Authenticate": "Bearer"},
        )
    except ContaIndisponivel:
        # 503, e não 401: o front derruba a sessão em todo 401
        # (`static/ui-common.js`), e uma instabilidade do banco não deve
        # deslogar quem está trabalhando.
        raise HTTPException(
            status_code=503,
            detail="Não foi possível confirmar a sua sessão. Tente novamente.",
        )

    if conta is None:
        return token_data
    if not auth.conta_ativa(conta):
        raise HTTPException(
            status_code=401,
            detail=SESSAO_ENCERRADA,
            headers={"WWW-Authenticate": "Bearer"},
        )
    if _senha_trocada_depois_do_token(conta, token_data):
        # Trocar a senha derruba as sessões abertas. Sem isto, quem redefine a
        # senha por suspeitar que alguém entrou continuaria com esse alguém
        # dentro por até 24h — exatamente enquanto acredita ter resolvido.
        raise HTTPException(
            status_code=401,
            detail=SESSAO_ENCERRADA,
            headers={"WWW-Authenticate": "Bearer"},
        )
    # Reconstruído a partir da linha, nunca do token: é o ponto inteiro daqui.
    # Com `.get`, e não `[...]`: papel ausente vira None, que `require_admin`
    # reprova. Indexar levantaria KeyError, e a rota responderia 500 — falha
    # aberta na cara do usuário onde cabia uma recusa silenciosa.
    return auth.TokenData(
        email=conta.get("email") or token_data.email,
        role=conta.get("role"),
        user_id=str(conta["id"]) if conta.get("id") is not None else None,
        emitido_em=token_data.emitido_em,
        deve_trocar_senha=bool(conta.get("deve_trocar_senha")),
    )


def _user_id_da_sessao(token: auth.TokenData) -> Optional[str]:
    """
    O UUID de quem está na requisição.

    `_sessao_do_token` já leu a linha inteira em `users` para validar a sessão, e
    o `id` veio junto. Reaproveitá-lo poupa repetir a MESMA consulta poucas
    linhas depois — foi o que pagou a leitura extra que a revogação introduziu.

    O `_resolve_user_id` continua como saída para quando não houve leitura: sem
    Supabase configurado, e no agente por X-API-Key. Nesses casos ele devolve o
    mesmo `None` de antes, e as rotas seguem respondendo 404 como respondiam.
    """
    return token.user_id or _resolve_user_id(token.email or "")


async def require_auth(
    request: Request,
    auth_creds: Optional[HTTPAuthorizationCredentials] = Depends(security),
    x_api_key: Optional[str] = Header(None, alias="X-API-Key"),
) -> auth.TokenData:
    """
    Dependência híbrida:
    1. Se houver Token JWT (Navegador), valida o usuário.
    2. Se houver X-API-Key (Agente Windows), valida a chave estática.
    """
    # 1. Tentar JWT. Assinatura válida ainda não é sessão válida: `_sessao_do_token`
    #    confere no banco se a conta segue existindo, ativa, e com qual papel.
    if auth_creds:
        token_data = auth.decode_access_token(auth_creds.credentials)
        if token_data:
            sessao = _sessao_do_token(token_data)
            # A barreira mora AQUI, e não na tela. Um modal pode ser fechado
            # pelo Esc, pelo devtools, ou simplesmente ignorado por quem chama
            # a API direto — e a senha provisória é conhecida por outra pessoa.
            if sessao.deve_trocar_senha and request.url.path not in ROTAS_COM_SENHA_PROVISORIA:
                raise HTTPException(
                    status_code=403,
                    detail=ERRO_SENHA_PROVISORIA,
                    # Cabeçalho, e não texto: o front precisa reconhecer este
                    # caso entre todos os 403 possíveis, e casar por string de
                    # mensagem quebraria ao reescrever a frase.
                    headers={"X-Senha-Provisoria": "1"},
                )
            return sessao

    # 2. Se API_KEY estiver ativa, aceitar a chave estática para o agente.
    if config.API_KEY:
        if x_api_key and x_api_key == config.API_KEY:
            return auth.TokenData(email="agent@internal", role="agent")
    else:
        # Ambiente aberto (sem API_KEY): mantém compatibilidade para rotas /api/*
        # que usam require_auth, sem elevar privilégios administrativos.
        # ATENÇÃO: esta identidade é anônima. Rotas que manipulam material
        # criptográfico devem recusá-la — ver require_agent_or_admin.
        return auth.TokenData(email=ANONYMOUS_IDENTITY_EMAIL, role="agent")

    raise HTTPException(
        status_code=401, 
        detail="Não autorizado. Faça login ou forneça uma chave de API válida.",
        headers={"WWW-Authenticate": "Bearer"},
    )

# Definidos em `auth` porque o login e o envio de alertas precisam da mesma
# regra; duas cópias divergiriam, e a divergência permissiva não daria sintoma.
PAPEIS_VALIDOS = auth.PAPEIS_VALIDOS
conta_ativa = auth.conta_ativa


async def require_admin(token: auth.TokenData = Depends(require_auth)) -> auth.TokenData:
    if token.role != "admin":
        raise HTTPException(status_code=403, detail="Acesso restrito a administradores.")
    return token


ERRO_ACESSO_MAQUINA = "Acesso restrito ao agente e a administradores."


def require_modulo(modulo: str, minimo: str = permissoes.NIVEL_LER):
    """
    Guarda por MODULO, lida da matriz de permissoes (`app/permissoes.py`).

    Substitui `require_admin` onde a alcada deixa de ser "so admin" e passa a
    ser configuravel pela tela. O nome do modulo e conferido AQUI, na
    importacao: um erro de digitacao vira erro de partida do servidor, e nao um
    403 silencioso em producao para um modulo que ninguem mexeu.

    NAO SERVE para rota que o AGENTE chama. O agente autentica por X-API-Key e
    recebe papel 'agent', que nao esta na matriz e portanto cai em `nenhum`.
    `GET /api/settings` e o exemplo vivo: pertence ao modulo `configuracao`, mas
    `agent/run_agent.py` o consome para saber quais pastas monitorar — liga-lo
    aqui pararia o agente em producao. Rota de maquina fica com
    `require_agent_or_admin`.
    """
    if modulo not in permissoes.MODULOS:
        raise ValueError(f"modulo desconhecido em require_modulo: {modulo!r}")
    if minimo not in permissoes.NIVEIS:
        raise ValueError(f"nivel desconhecido em require_modulo: {minimo!r}")

    async def _guarda(token: auth.TokenData = Depends(require_auth)) -> auth.TokenData:
        try:
            if permissoes.pode(token.role or "", modulo, minimo):
                return token
        except permissoes.PermissoesIndisponiveis:
            # 503, e nao 403: "nao consegui verificar" nao e "voce nao pode".
            # Mesmo criterio de `require_admin_ou_lider` e `_exigir_alcance`.
            raise HTTPException(
                status_code=503,
                detail="Nao foi possivel verificar suas permissoes. Tente de novo.",
            )
        raise HTTPException(
            status_code=403,
            detail=f"Seu perfil nao tem acesso a {modulo}.",
        )

    return _guarda


async def require_agent_or_admin(token: auth.TokenData = Depends(require_auth)) -> auth.TokenData:
    """
    Endpoints da máquina: alimentados pelo agente (X-API-Key -> role 'agent') e
    acessíveis a administradores para diagnóstico.

    Existe porque `require_auth` sozinho é permissivo demais para estas rotas.

    1. Aceita qualquer usuário do portal, inclusive role 'user'. Em /upload-pfx
       isso permitia a um usuário comum enviar um PFX próprio reaproveitando o
       fingerprint de um certificado já armazenado: como `upsert_pfx` usa
       `on_conflict="fingerprint"`, o registro legítimo seria sobrescrito e o
       certificado do atacante acabaria instalado num servidor pelo fluxo
       normal de instalação. `require_admin` não serve como alternativa: o
       agente tem role 'agent', não 'admin'.

    2. Quando API_KEY não está configurada, `require_auth` devolve uma
       identidade ANÔNIMA com role 'agent' para manter compatibilidade. Para as
       rotas antigas isso é aceitável; para estas, que entregam PFX e senhas,
       significaria acesso sem credencial nenhuma. Aqui a identidade anônima é
       recusada explicitamente.
    """
    if token.role not in ("agent", "admin") or token.email == ANONYMOUS_IDENTITY_EMAIL:
        raise HTTPException(status_code=403, detail=ERRO_ACESSO_MAQUINA)
    return token

class SecureJSONFormatter(logging.Formatter):
    def format(self, record: logging.LogRecord) -> str:
        # [OWASP A09] Ocultação de segredos e geração de Log JSON estruturado para prevenir Log Injection
        log_obj = {
            # O replace preserva o sufixo "Z": isoformat() de um datetime aware
            # emite "+00:00", e concatenar "Z" daria "+00:00Z". O formato do log
            # continua byte a byte igual ao de antes.
            "timestamp": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "level": record.levelname,
            "logger": record.name,
            "message": record.getMessage(),
        }
        # Filtrar possíveis senhas ou tokens da mensagem bruta
        msg_lower = log_obj["message"].lower()
        if "password" in msg_lower or "token" in msg_lower or "senha" in msg_lower:
            log_obj["message"] = "*** REDACTED SENSITIVE DATA ***"
            
        if record.exc_info:
            log_obj["exc_info"] = self.formatException(record.exc_info)
        return json.dumps(log_obj)

logger = logging.getLogger(__name__)
# Configuração base de logging
_handler = logging.StreamHandler()
_handler.setFormatter(SecureJSONFormatter())
logging.root.handlers = [_handler]
logging.root.setLevel(logging.INFO)

LISTAGEM_EXPORT_MAX = 5000

@asynccontextmanager
async def lifespan(_app: FastAPI):
    """
    Inicialização e encerramento do portal.

    Substitui `@app.on_event("startup")`, deprecado no FastAPI. A ordem é a
    mesma de antes: criar os diretórios locais e só então subir o job de
    alertas.

    O que muda além da deprecação: o `asyncio.create_task` do laço de alertas
    não tinha contrapartida no shutdown — a task ficava pendurada e, em
    reinícios rápidos, um novo laço subia enquanto o anterior ainda dormia. O
    `finally` agora cancela e aguarda o encerramento.
    """
    import asyncio

    # Antes de qualquer outra coisa: sem a chave de cifragem da senha SMTP o
    # portal não sobe. Deliberadamente fatal — o desenho anterior derivava uma
    # chave da JWT_SECRET_KEY e seguia em frente, e a consequência (senha SMTP
    # indecifrável) só aparecia como "os alertas pararam", desligado no tempo e
    # no espaço da causa.
    #
    # ATENÇÃO AO DEPLOY: defina ENCRYPTION_KEY no painel da plataforma
    # (Vercel/Render) ANTES de publicar esta versão. O .env não sobe no deploy.
    smtp_service.verificar_chave_configurada()

    try:
        config.CERT_SOURCE_DIR.mkdir(parents=True, exist_ok=True)
        config.CERT_EXPIRED_DIR.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        logger.warning(f"Não foi possível criar diretórios locais (ambiente read-only / Vercel): {e}")

    # Job diário de alertas por e-mail.
    #
    # Dois problemas do desenho anterior:
    #
    # 1. `trigger_all_alerts()` é síncrona e era chamada direto no laço async.
    #    Ela faz a varredura completa dos certificados (2,5-3,3s numa base de
    #    mil) e depois N envios SMTP sequenciais — tudo isso travava o event
    #    loop, ou seja, o portal inteiro parava de responder durante o job.
    #    Agora roda em thread separada via run_in_executor.
    #
    # 2. O laço dormia 86400s, mas o Procfile usa `--max-requests 500`: o worker
    #    recicla várias vezes ao dia e o job redisparava em cada boot+60s.
    #    "Diário" nunca foi diário. O marcador em disco (job_ja_executado_
    #    recentemente) torna a cadência real, independente de reinícios.
    async def daily_alerts_job_loop():
        logger.info("Iniciando loop do job diário de alertas")
        await asyncio.sleep(60)  # Deixa o boot terminar antes do primeiro disparo
        loop = asyncio.get_running_loop()
        while True:
            try:
                if job_ja_executado_recentemente():
                    logger.info("Job de alertas ignorado: já executado nas últimas horas.")
                else:
                    logger.info("Executando job de alertas por e-mail...")
                    stats = await loop.run_in_executor(None, trigger_all_alerts)
                    logger.info(f"Job de alertas concluído: {stats}")
            except Exception as e:
                logger.error(f"Erro ao executar job de alertas: {e}")
            # Reavalia de hora em hora: com o marcador de última execução, o
            # trabalho real acontece uma vez por dia mesmo com ciclo curto.
            await asyncio.sleep(3600)

    tarefa_alertas = asyncio.create_task(daily_alerts_job_loop())
    try:
        yield
    finally:
        # O `await` é limitado no tempo de propósito. Esperar a task sem prazo
        # significa que qualquer falha em encerrá-la (um cancel que não chega,
        # um run_in_executor preso num envio SMTP) trava o shutdown do processo
        # para sempre — o worker não recicla e o deploy não termina. Desistir
        # depois de alguns segundos e registrar é melhor que pendurar.
        # `asyncio.wait` em vez de `wait_for` + `except CancelledError`: aquele
        # except engolia qualquer cancelamento, inclusive um dirigido ao próprio
        # lifespan por quem o encerra — o shutdown virava inignorável e a espera
        # pelo prazo cheio passava despercebida. `asyncio.wait` devolve a task
        # pendente sem levantar, e deixa passar um cancelamento externo.
        tarefa_alertas.cancel()
        _, pendentes = await asyncio.wait({tarefa_alertas}, timeout=5)
        if pendentes:
            logger.warning("Job de alertas não encerrou em 5s; seguindo com o shutdown.")


app = FastAPI(title="Monitor de certificados PFX", version="1.2.1", lifespan=lifespan)

import secrets

@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    # Gera um nonce único por requisição para blindar scripts
    nonce = secrets.token_urlsafe(16)
    request.state.nonce = nonce
    
    response = await call_next(request)
    
    # [Guia Definitivo - A+ / Mozilla Observatory] Headers Críticos
    response.headers["Strict-Transport-Security"] = "max-age=63072000; includeSubDomains; preload"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=(), payment=(), usb=()"
    
    # [Isolamento de Origem Cruzada - Mitigação Spectre]
    response.headers["Cross-Origin-Resource-Policy"] = "same-origin"
    response.headers["Cross-Origin-Opener-Policy"] = "same-origin"
    
    # Permissão dev-only para o modo live do Impeccable, cujo picker de variantes
    # é servido em http://localhost:8400. Só existe quando IMPECCABLE_LIVE=1 está
    # no ambiente; a Vercel nunca define essa variável, então o header de produção
    # sai byte a byte igual ao de antes desta linha existir.
    _live = " http://localhost:8400" if os.getenv("IMPECCABLE_LIVE") == "1" else ""

    # CSP Avançado (Removido unsafe-inline/unsafe-eval de script-src e adicionado nonce)
    csp = (
        "default-src 'self'; "
        f"script-src 'self' 'nonce-{nonce}'{_live}; "
        # `connect-src` herdava de `default-src 'self'`. Declarado explicitamente
        # com o mesmo valor para receber a permissão dev acima — com `_live`
        # vazio, herdar e declarar dão exatamente o mesmo resultado.
        f"connect-src 'self'{_live}; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "object-src 'none'; "
        "form-action 'self'; "
        "frame-ancestors 'none'; "
        "block-all-mixed-content; "
        "upgrade-insecure-requests;"
    )
    response.headers["Content-Security-Policy"] = csp
    
    # Remoção de cabeçalhos de rastreamento/obsoletos
    if "X-XSS-Protection" in response.headers:
        del response.headers["X-XSS-Protection"]
    if "Server" in response.headers:
        del response.headers["Server"]
    if "X-Powered-By" in response.headers:
        del response.headers["X-Powered-By"]
        
    return response

templates = Jinja2Templates(directory=str(ROOT / "templates"))
app.mount("/static", StaticFiles(directory=str(ROOT / "static")), name="static")


# As funções require_api_key foram removidas em favor do require_auth híbrido.


def _painel_busca_normalizada(value: Any) -> str:
    """Mesma ideia que `normalizarTexto` no dashboard (minúsculas, sem acentos)."""
    t = str(value or "").lower()
    t = unicodedata.normalize("NFD", t)
    return "".join(ch for ch in t if unicodedata.category(ch) != "Mn")


def _enrich_cert_item_dashboard_flags(it: dict, now: datetime, thirty_days: datetime) -> dict:
    """Alinha com o painel: vencido pela data `not_after` mesmo que `status` ainda seja ok."""
    row = dict(it)
    expired_by_date = False
    expiring_soon = False
    na = row.get("not_after")
    if na:
        exp = _parse_iso_utc(str(na))
        min_dt = datetime.min.replace(tzinfo=timezone.utc)
        if exp > min_dt:
            if exp < now:
                expired_by_date = True
            elif exp <= thirty_days and exp >= now:
                expiring_soon = True
    row["_isExpiredByDate"] = expired_by_date
    row["_isExpiringSoon"] = expiring_soon
    return row


def _dashboard_filtro_status_match(row: dict, filtro: str) -> bool:
    s = str(row.get("status") or "").lower()
    ed = bool(row.get("_isExpiredByDate"))
    es = bool(row.get("_isExpiringSoon"))
    f = (filtro or "todos").strip().lower()
    if f in ("todos", ""):
        return True
    if f == "validos":
        return s == "ok" and not ed and not es
    if f in ("prestes_vencer", "prestes a vencer"):
        return s == "ok" and es and not ed
    if f == "vencidos":
        return s == "expirado" or ed
    if f in ("erros", "erro"):
        return s == "erro"
    if f in ("sem_padrao", "fora_do_padrao"):
        return s == "fora_do_padrao"
    return True


def _dashboard_busca_match(row: dict, q_raw: str) -> bool:
    if not str(q_raw or "").strip():
        return True
    raw = str(q_raw).strip()
    bt = _painel_busca_normalizada(raw)
    bd = re.sub(r"\D", "", raw)
    nome = _painel_busca_normalizada(row.get("nome") or row.get("display_name") or "")
    doc_f = _painel_busca_normalizada(row.get("documento_formatado") or "")
    doc_n = _painel_busca_normalizada(row.get("documento_numero") or "")
    fn = _painel_busca_normalizada(row.get("file_name") or "")
    na_txt = _painel_busca_normalizada(row.get("not_after") or "")
    dd = _digits_only_doc(row.get("documento_numero") or row.get("documento_formatado"))
    nd = _digits_only_doc(str(row.get("not_after") or ""))
    if bt and bt in nome:
        return True
    if bt and bt in doc_f:
        return True
    if bt and bt in doc_n:
        return True
    if bt and bt in fn:
        return True
    if bt and bt in na_txt:
        return True
    if bd and bd in dd:
        return True
    if bd and bd in nd:
        return True
    return False


def _dashboard_resumo_counts(rows: List[dict]) -> dict[str, int]:
    total = len(rows)
    validos = expirando = erros = vencidos = sem_padrao = 0
    for it in rows:
        s = str(it.get("status") or "").lower()
        ed = bool(it.get("_isExpiredByDate"))
        es = bool(it.get("_isExpiringSoon"))
        if s == "erro":
            erros += 1
        elif s == "fora_do_padrao":
            sem_padrao += 1
        elif s == "expirado" or ed:
            vencidos += 1
        elif s == "ok":
            if es:
                expirando += 1
            else:
                validos += 1
    return {
        "total": total,
        "validos": validos,
        "expirando": expirando,
        "erros": erros,
        "vencidos": vencidos,
        "sem_padrao": sem_padrao,
    }


def chave_alfabetica(item: dict) -> tuple:
    """
    Chave de ordenação por titular, insensível a caixa e acento.

    `(1, "")` para quem não tem nome legível: vai para o fim. Ordenar por
    string vazia os jogaria para o topo, e a primeira página da lista seria
    justamente o que o robô não conseguiu ler — o oposto do útil.
    """
    bruto = str(item.get("nome") or item.get("display_name") or item.get("file_name") or "").strip()
    if not bruto:
        return (1, "")
    sem_acento = "".join(
        c for c in unicodedata.normalize("NFD", bruto) if unicodedata.category(c) != "Mn"
    )
    return (0, sem_acento.casefold())


def ordenar_por_titular(itens: List[dict]) -> List[dict]:
    """
    Ordena a listagem de certificados pelo nome do titular.

    **Precisa acontecer aqui, no servidor, e antes da paginação.** A tabela
    pagina no servidor: ordenar no navegador ordenaria só os 10 ou 25 itens da
    página visível, e a lista *pareceria* certa enquanto continuasse errada
    entre páginas — que é pior que estar visivelmente errada.

    Não havia ordenação nenhuma antes. A lista saía na ordem em que o agente
    varre o disco, e ele percorre pasta por pasta: o resultado eram vários
    blocos alfabéticos emendados (um por subpasta, mais os vencidos no fim),
    que de longe parece ordem alfabética e de perto não é.
    """
    return sorted(itens, key=chave_alfabetica)


def _list_certificados_payload(
    sets: PortalSettings,
    snap: Optional[dict],
    fonte: str,
) -> dict[str, Any]:
    """Monta o payload base (sem paginação) para GET /api/certificados."""
    if fonte == "local":
        src = sets.effective_source()
        exp = sets.effective_expired()
        itens: List[CertInfo] = scan_folder(src)
        return {
            "source_dir": str(src),
            "expired_dir": str(exp),
            "atualizado_em": datetime.now(timezone.utc).isoformat(),
            "itens": [cert_to_public_dict(c) for c in itens],
            "data_source": "local",
            "machine_id": sets.machine_id,
        }
    if fonte == "remoto":
        if not snap:
            raise HTTPException(
                status_code=404,
                detail="Nenhum dado remoto. Configure o agente no Windows para enviar leituras.",
            )
        return {
            "source_dir": str(snap.get("source_folder", "") or ""),
            "expired_dir": str(snap.get("expired_folder", "") or ""),
            "atualizado_em": snap.get("scanned_at", datetime.now(timezone.utc).isoformat()),
            "itens": list(snap.get("items", []) or []),
            "data_source": "remoto",
            "machine_id": snap.get("machine_id"),
        }
    # auto
    if snap:
        return {
            "source_dir": str(snap.get("source_folder", "") or ""),
            "expired_dir": str(snap.get("expired_folder", "") or ""),
            "atualizado_em": snap.get("scanned_at", datetime.now(timezone.utc).isoformat()),
            "itens": list(snap.get("items", []) or []),
            "data_source": "remoto",
            "machine_id": snap.get("machine_id"),
        }
    src = sets.effective_source()
    exp = sets.effective_expired()
    itens_scan: List[CertInfo] = scan_folder(src)
    return {
        "source_dir": str(src),
        "expired_dir": str(exp),
        "atualizado_em": datetime.now(timezone.utc).isoformat(),
        "itens": [cert_to_public_dict(c) for c in itens_scan],
        "data_source": "local",
        "machine_id": sets.machine_id,
    }


class SettingsBody(BaseModel):
    source_folder: str = Field(default="", description="Pasta de certificados no Windows (caminho completo)")
    expired_folder: str = Field(default="", description="Pasta destino dos vencidos")
    machine_id: str = Field(default="default", description="Identificador lógico da máquina / agente")
    smtp_host: str = Field(default="")
    smtp_port: int = Field(default=587)
    smtp_user: str = Field(default="")
    smtp_password: Optional[str] = Field(default=None)
    smtp_use_tls: bool = Field(default=True)
    smtp_use_ssl: bool = Field(default=False)
    smtp_from_email: str = Field(default="")
    smtp_alerts_enabled: bool = Field(default=False)
    # Módulo instalador. Vazio/zero = usar o padrão do código, e não "desligado":
    # uma configuração nunca tocada tem de se comportar como antes de existir.
    instalador_nome_template: str = Field(default="")
    install_token_ttl_min: int = Field(default=0)
    trilha_retencao_dias: int = Field(default=0)


class IngestBody(BaseModel):
    machine_id: str = "default"
    source_folder: str
    expired_folder: str
    items: List[dict] = Field(default_factory=list)
    scanned_at: Optional[str] = None


class EnqueueCommandBody(BaseModel):
    machine_id: str = "default"
    command: str = Field(..., description="mover_vencidos | rescan | ping")


# A `pagina_ativa` acende o item correspondente em templates/_sidebar.html.
# Rota que esquecer de passa-la renderiza o menu sem nenhum item aceso - falha
# silenciosa, por isso `tests/test_sidebar_partial.py` cobre todas elas.
@app.get("/", response_class=HTMLResponse)
def painel(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={
            "pagina_ativa": "inicio",
            # O teto vem do servidor para não existir em dois lugares. Um número
            # digitado na tela divergiria do que a rota aceita, e o sintoma seria
            # o pior dos dois: a tela deixa marcar 60, o download falha com 422.
            "max_certificados": MAX_CERTIFICADOS_POR_TOKEN,
        },
    )


@app.get("/configuracao", response_class=HTMLResponse)
def pagina_configuracao(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request=request, name="configuracao.html", context={"pagina_ativa": "configuracao"}
    )


@app.get("/login", response_class=HTMLResponse)
def pagina_login(request: Request) -> HTMLResponse:
    # Sem sidebar: quem nao entrou ainda nao tem para onde navegar.
    return templates.TemplateResponse(request=request, name="login.html")


@app.get("/usuarios", response_class=HTMLResponse)
def pagina_usuarios(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request=request, name="usuarios.html", context={"pagina_ativa": "usuarios"}
    )


@app.get("/historico", response_class=HTMLResponse)
def pagina_historico(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request=request, name="historico.html", context={"pagina_ativa": "historico"}
    )


@app.get("/vencidos", response_class=HTMLResponse)
def pagina_vencidos(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request=request, name="vencidos.html", context={"pagina_ativa": "vencidos"}
    )


@app.get("/duplicidades", response_class=HTMLResponse)
def pagina_duplicidades(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request=request, name="duplicidades.html", context={"pagina_ativa": "duplicidades"}
    )


@app.get("/acompanhamento", response_class=HTMLResponse)
def pagina_colaborador_certificados(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request=request, name="colaborador_certificados.html", context={"pagina_ativa": "acompanhamento"}
    )


@app.get("/favicon.ico", include_in_schema=False)
def favicon() -> Response:
    """
    Serve o favicon do projeto quando disponível.
    Fallback 204 para evitar ruído de 404 no log em dev.
    """
    icon_path = ROOT / "ico" / "icone.ico"
    if icon_path.is_file():
        return FileResponse(path=icon_path, media_type="image/x-icon")
    return Response(status_code=204)


class LoginBody(BaseModel):
    email: str
    password: str


@app.post("/api/login")
def login(body: LoginBody, request: Request) -> dict:
    load_settings()  # trigger client init
    from app.settings_state import _supabase
    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Sistema sem Supabase configurado para login.")

    ip = request.client.host if request and request.client else None
    # Normalizado ANTES da consulta. A coluna guarda tudo em minúsculas (o
    # portal grava assim, e o índice único é sobre `lower(email)`), mas a
    # consulta usava o valor cru do formulário: quem digitasse "Ana@X.com" não
    # casava com linha nenhuma e levava 401 "E-mail ou senha incorretos" —
    # mensagem que manda a pessoa caçar a senha por causa de uma maiúscula.
    #
    # O `trim` está aqui pelo mesmo motivo: colar o e-mail de um e-mail
    # costuma trazer espaço junto.
    email_login = (body.email or "").strip().lower()
    try:
        r = sb.table("users").select("*").eq("email", email_login).limit(1).execute()
        user = r.data[0] if r.data else None
        if not user or not auth.verify_password(body.password, user["password_hash"]):
            # Só registra quando a conta EXISTE: e-mail inexistente viraria
            # guardar entrada arbitrária de quem chamou. Com conta existente, o
            # registro responde "alguém está tentando entrar aqui", que é o
            # caso que importa.
            if user:
                atividade.registrar(
                    atividade.EVENTO_LOGIN_NEGADO,
                    user_id=str(user.get("id") or "") or None,
                    user_email=email_login,
                    client_ip=ip,
                    contexto={"motivo": "senha_incorreta"},
                )
            raise HTTPException(status_code=401, detail="E-mail ou senha incorretos.")
        if not conta_ativa(user):
            atividade.registrar(
                atividade.EVENTO_LOGIN_NEGADO,
                user_id=str(user.get("id") or "") or None,
                user_email=email_login,
                client_ip=ip,
                contexto={"motivo": "conta_desativada"},
            )
            raise HTTPException(status_code=403, detail="Usuário desativado. Procure um administrador.")

        atividade.registrar(
            atividade.EVENTO_LOGIN,
            user_id=str(user.get("id") or "") or None,
            user_email=user["email"],
            client_ip=ip,
        )
        token = auth.create_access_token({"sub": user["email"], "role": user["role"]})
        return {"access_token": token, "token_type": "bearer", "role": user["role"]}
    except HTTPException:
        # O `except Exception` abaixo engolia estas: uma senha errada saía como
        # 500 com "401: E-mail ou senha incorretos." no corpo — mensagem certa,
        # status errado, e todo tratamento no front que olhasse o código via
        # "erro do servidor" onde houve credencial inválida.
        raise
    except Exception as e:
        logger.exception("Erro no login")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/users", dependencies=[Depends(require_admin)])
def list_users() -> List[dict]:
    from app.settings_state import _supabase
    sb = _supabase()
    if not sb: return []
    r = sb.table("users").select(
        "id, email, full_name, role, ativo, gestor_id, departamento_id, created_at"
    ).execute()
    return r.data


class UserCreateBody(BaseModel):
    email: str
    password: str
    full_name: str
    role: str = "user"
    departamento_id: Optional[str] = None


class UserUpdateBody(BaseModel):
    email: str
    full_name: str
    role: str = "user"
    # Omitir mantém o que está gravado. `role: "disabled"` continua aceito como
    # forma antiga de desativar (ver `update_user`), mas não escreve mais em
    # `role` — o papel deixou de ser o lugar onde o estado mora.
    ativo: Optional[bool] = None
    gestor_id: Optional[str] = None
    # Omitir mantém o que está gravado; string vazia limpa. Sem a distinção,
    # não haveria como tirar alguém de um setor sem inventar um valor.
    departamento_id: Optional[str] = None


class UserResetPasswordBody(BaseModel):
    password: str


def _norm_header(v: str) -> str:
    s = unicodedata.normalize("NFD", str(v or "").strip().lower())
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return s


@app.post("/api/users/import", dependencies=[Depends(require_admin)])
async def import_users(file: UploadFile = File(...)) -> dict:
    from app.settings_state import _supabase

    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Sistema sem Supabase configurado.")

    name = (file.filename or "").lower()
    if not name.endswith(".csv"):
        raise HTTPException(
            status_code=422,
            detail="Formato inválido. Exporte a planilha como CSV e envie um arquivo .csv.",
        )

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=422, detail="Arquivo vazio.")
        
    # [OWASP A08] Validação de Limite de Tamanho
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Arquivo muito grande (limite de 5MB).")
        
    # [OWASP A08] Validação de Magic Bytes (Assinatura real do arquivo)
    # Rejeita ativamente se for um binário executável ou arquivo restrito disfarçado de CSV
    if raw.startswith(b'MZ') or raw.startswith(b'\x7fELF') or raw.startswith(b'%PDF') or raw.startswith(b'PK'):
        raise HTTPException(status_code=422, detail="Conteúdo do arquivo suspeito. Apenas texto puro (CSV) é permitido.")

    text = raw.decode("utf-8-sig", errors="replace")
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:2048], delimiters=",;")
        delim = dialect.delimiter
    except csv.Error:
        delim = ";"

    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    if not reader.fieldnames:
        raise HTTPException(status_code=422, detail="CSV sem cabeçalho.")

    map_headers = {_norm_header(h): h for h in reader.fieldnames}

    def pick(*aliases: str) -> Optional[str]:
        for a in aliases:
            key = map_headers.get(_norm_header(a))
            if key:
                return key
        return None

    h_nome = pick("nome", "full_name", "nome completo")
    h_email = pick("email", "e-mail")
    h_senha = pick("senha", "password")
    h_role = pick("role", "nivel", "papel", "perfil")
    if not h_nome or not h_email or not h_senha:
        raise HTTPException(
            status_code=422,
            detail="Cabeçalho obrigatório: nome, email, senha.",
        )
    if not h_role:
        raise HTTPException(
            status_code=422,
            detail="Cabeçalho obrigatório também para nível: use 'nivel' ou 'role' com valores 'admin' ou 'user'.",
        )

    criados = 0
    ignorados = 0
    erros: List[dict[str, Any]] = []
    linha = 1
    for row in reader:
        linha += 1
        nome = str(row.get(h_nome) or "").strip()
        email = str(row.get(h_email) or "").strip().lower()
        senha = str(row.get(h_senha) or "").strip()
        role = str(row.get(h_role) or "").strip().lower()

        if not nome or not email or not senha or not role:
            ignorados += 1
            continue
        if role not in PAPEIS_VALIDOS:
            erros.append(
                {
                    "linha": linha,
                    "email": email,
                    "erro": f"Nível inválido. Use exatamente: {', '.join(PAPEIS_VALIDOS)}.",
                }
            )
            continue
        if len(senha) < SENHA_MINIMA:
            erros.append({"linha": linha, "email": email,
                          "erro": f"Senha deve ter no mínimo {SENHA_MINIMA} caracteres."})
            continue
        # O CSV era o caminho que escapava de tudo: nem o formulário HTML o
        # cobre, nem a API validava.
        if not _EMAIL_PLAUSIVEL.match(email):
            erros.append({"linha": linha, "email": email, "erro": "E-mail inválido."})
            continue
        try:
            existe = sb.table("users").select("id").eq("email", email).limit(1).execute()
            if existe.data:
                ignorados += 1
                continue
            sb.table("users").insert(
                {
                    "email": email,
                    "password_hash": auth.get_password_hash(senha),
                    "full_name": nome,
                    "role": role,
                }
            ).execute()
            criados += 1
        except Exception as e:  # noqa: BLE001
            erros.append({"linha": linha, "email": email, "erro": str(e)})

    return {"ok": True, "criados": criados, "ignorados": ignorados, "erros": erros}


# Validação deliberadamente frouxa: exige um "@" com algo dos dois lados e um
# ponto no domínio, e nada além disso. Regex de e-mail "completo" rejeita
# endereços válidos e dá falsa sensação de rigor — o que prova que um e-mail
# funciona é uma mensagem chegar nele.
_EMAIL_PLAUSIVEL = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]{2,}$")

# O mesmo mínimo que `reset_user_password` já exigia. Antes, criar era mais
# permissivo que redefinir: dava para nascer com senha de um caractere e só
# descobrir o rigor ao trocá-la.
SENHA_MINIMA = 6


def _validar_email(email: str) -> str:
    limpo = (email or "").strip().lower()
    if not _EMAIL_PLAUSIVEL.match(limpo):
        raise HTTPException(status_code=422, detail=f"E-mail inválido: {email!r}")
    return limpo


def _garantir_email_livre(sb: Any, email: str, ignorar_id: Optional[str] = None) -> None:
    """
    Recusa e-mail já usado por outra conta.

    O índice único no banco é quem fecha de verdade — esta checagem tem janela
    de corrida e existe para a mensagem ser legível em vez de um 400 cru do
    PostgREST. As duas camadas servem a coisas diferentes.
    """
    try:
        existentes = sb.table("users").select("id, email").execute().data or []
    except Exception:
        logger.exception("Falha ao verificar e-mail duplicado")
        raise HTTPException(
            status_code=503, detail="Não foi possível verificar o e-mail agora."
        )
    for u in existentes:
        if str(u.get("email") or "").strip().lower() == email and str(u.get("id")) != str(ignorar_id):
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Já existe uma conta com o e-mail {email}. O e-mail identifica "
                    "a pessoa no login — duas contas com o mesmo endereço deixam "
                    "uma delas inacessível."
                ),
            )


@app.post("/api/users", dependencies=[Depends(require_admin)])
def create_user(body: UserCreateBody) -> dict:
    from app.settings_state import _supabase
    sb = _supabase()
    if not sb: raise HTTPException(status_code=503)
    
    # A validação não existia aqui: qualquer string virava papel. Com o CHECK
    # no banco isso passaria a estourar como 400 genérico do PostgREST, sem
    # dizer qual valor era aceito.
    role = (body.role or "user").strip().lower()
    if role not in PAPEIS_VALIDOS:
        raise HTTPException(
            status_code=422,
            detail=f"Nível inválido. Use: {', '.join(PAPEIS_VALIDOS)}.",
        )

    email = _validar_email(body.email)
    if len((body.password or "").strip()) < SENHA_MINIMA:
        raise HTTPException(
            status_code=422,
            detail=f"A senha precisa ter no mínimo {SENHA_MINIMA} caracteres.",
        )
    _garantir_email_livre(sb, email)

    hash_pw = auth.get_password_hash(body.password)
    try:
        sb.table("users").insert({
            # Quem cadastrou sabe a senha que digitou. Ela serve para o primeiro
            # acesso e nada mais — `require_auth` recusa o resto do portal até a
            # pessoa escolher uma própria.
            "deve_trocar_senha": True,
            "departamento_id": (body.departamento_id or "").strip() or None,
            "email": email,
            "password_hash": hash_pw,
            "full_name": body.full_name,
            "role": role,
            "ativo": True,
        }).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


def _garantir_que_sobra_admin(
    sb: Any,
    user_id: str,
    *,
    novo_role: Optional[str] = None,
    novo_ativo: Optional[bool] = None,
    apagar: bool = False,
) -> None:
    """
    Recusa a operação se ela deixaria o portal **sem nenhum administrador ativo**.

    A regra é "tem de sobrar um", e não "não mexa em si mesmo". A segunda
    formulação parece equivalente e não é: com dois admins, um poderia
    rebaixar o outro e depois sair, e nenhuma das duas ações seria sobre si
    mesmo. E com um admin só — que é o caso do portal hoje — a versão correta
    também bloqueia desativar, rebaixar e apagar, que são três caminhos para o
    mesmo buraco.

    O buraco é sem fundo: a tela de Usuários é `require_admin`, então sem admin
    ativo **não há como voltar pela interface**. Só SQL direto no banco.

    Simula a mudança e conta o que restaria — assim as três rotas usam a mesma
    regra e não há como uma delas divergir.
    """
    try:
        us = sb.table("users").select("id, role, ativo").execute().data or []
    except Exception:
        # Não dá para afirmar que sobra admin. Recusar é o lado seguro: o custo
        # é uma operação adiada; o do contrário é um portal sem dono.
        logger.exception("Falha ao verificar administradores restantes")
        raise HTTPException(
            status_code=503,
            detail="Não foi possível verificar os administradores agora. Tente novamente.",
        )

    restantes = 0
    for u in us:
        if str(u.get("id")) == str(user_id):
            if apagar:
                continue
            papel = novo_role if novo_role is not None else (u.get("role") or "")
            ativo = novo_ativo if novo_ativo is not None else u.get("ativo")
            u = {**u, "role": papel, "ativo": ativo}
        if (u.get("role") or "").strip().lower() == "admin" and auth.conta_ativa(u):
            restantes += 1

    if restantes == 0:
        raise HTTPException(
            status_code=409,
            detail=(
                "Esta é a única conta de administrador ativa. Promova outro "
                "administrador antes de desativar, rebaixar ou apagar esta — "
                "sem nenhum admin, não há como voltar pela tela."
            ),
        )


@app.put("/api/users/{user_id}", dependencies=[Depends(require_admin)])
def update_user(user_id: str, body: UserUpdateBody) -> dict:
    from app.settings_state import _supabase
    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503)
    role = (body.role or "user").strip().lower()
    ativo = body.ativo

    # Cliente antigo mandando role="disabled" quer dizer "desative" — nunca
    # quis dizer "o papel dele agora é disabled", embora fosse isso que
    # acontecia. Traduz para o estado e preserva o papel gravado.
    if role == "disabled":
        ativo = False
        role = None

    if role is not None and role not in PAPEIS_VALIDOS:
        raise HTTPException(
            status_code=422,
            detail=f"Nível inválido. Use: {', '.join(PAPEIS_VALIDOS)}.",
        )

    email = _validar_email(body.email)
    _garantir_email_livre(sb, email, ignorar_id=user_id)
    _garantir_que_sobra_admin(sb, user_id, novo_role=role, novo_ativo=ativo)

    campos: Dict[str, Any] = {
        "email": email,
        "full_name": body.full_name.strip(),
    }
    if role is not None:
        campos["role"] = role
    if ativo is not None:
        campos["ativo"] = bool(ativo)
    if body.departamento_id is not None:
        did = body.departamento_id.strip()
        campos["departamento_id"] = did or None
    if body.gestor_id is not None:
        gid = body.gestor_id.strip()
        if gid and gid == user_id:
            raise HTTPException(status_code=422, detail="Um usuário não pode ser gestor de si mesmo.")
        campos["gestor_id"] = gid or None

    # Nada a fazer com as seleções de alerta ao trocar o e-mail: desde a fase
    # 3c elas são chaveadas por `user_id`, então a identidade não se move. O
    # `_mover_selecoes_de_email` que existia aqui, e a leitura do endereço
    # anterior que o alimentava, viraram código morto e saíram — poder apagar
    # aquele helper era o sinal de que o rechaveamento tinha terminado.
    try:
        sb.table("users").update(campos).eq("id", user_id).execute()
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True}


@app.post("/api/users/{user_id}/reset-password", dependencies=[Depends(require_admin)])
def reset_user_password(user_id: str, body: UserResetPasswordBody) -> dict:
    from app.settings_state import _supabase
    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503)
    new_pw = (body.password or "").strip()
    if len(new_pw) < 6:
        raise HTTPException(status_code=422, detail="Senha deve ter no mínimo 6 caracteres.")
    hash_pw = auth.get_password_hash(new_pw)
    try:
        # Mesma razão do cadastro: o admin conhece a senha que acabou de
        # digitar. Sem isto ela valeria indefinidamente.
        sb.table("users").update(
            {"password_hash": hash_pw, "deve_trocar_senha": True}
        ).eq("id", user_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/users/{user_id}/deactivate", dependencies=[Depends(require_admin)])
def deactivate_user(user_id: str) -> dict:
    """
    Desativa a conta **preservando o papel**.

    Até 15/08 isto gravava `role = "disabled"`, o que apagava o papel: reativar
    um administrador virava adivinhação, e o mesmo teria acontecido com gestor —
    levando junto o sentido das carteiras que ele tivesse criado.
    """
    from app.settings_state import _supabase
    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503)
    _garantir_que_sobra_admin(sb, user_id, novo_ativo=False)
    try:
        sb.table("users").update({"ativo": False}).eq("id", user_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/users/{user_id}/reactivate", dependencies=[Depends(require_admin)])
def reactivate_user(user_id: str) -> dict:
    """
    Reativa a conta, devolvendo o papel que ela sempre teve.

    Não existia contrapartida para o desativar: o único caminho de volta era
    editar o nível na mão e escolher um papel de memória. Com estado e papel
    separados, reativar deixa de ser uma decisão.

    As contas desativadas ANTES desta separação são a exceção: o papel delas foi
    sobrescrito e a migration as pôs em 'user', o menor privilégio. Se alguma
    era admin, promover é ato explícito — e é assim que deve ser.
    """
    from app.settings_state import _supabase
    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503)
    try:
        sb.table("users").update({"ativo": True}).eq("id", user_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.delete("/api/users/{user_id}", dependencies=[Depends(require_admin)])
def delete_user(user_id: str) -> dict:
    from app.settings_state import _supabase
    sb = _supabase()
    if not sb: raise HTTPException(status_code=503)
    _garantir_que_sobra_admin(sb, user_id, apagar=True)
    sb.table("users").delete().eq("id", user_id).execute()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════
# Departamentos
#
# O departamento recorta quem cada líder pode liberar. Criar, renomear e
# apagar setor é de ADMIN: quem define os setores define, por consequência, o
# alcance de cada líder — deixar isso com o próprio líder o deixaria ampliar o
# próprio alcance criando setores e se pondo neles.
# ══════════════════════════════════════════════════════════════════════════


class DepartamentoBody(BaseModel):
    nome: str


class DepartamentoLideresBody(BaseModel):
    lideres: List[str] = Field(default_factory=list)


def _nome_de_departamento(nome: str) -> str:
    limpo = (nome or "").strip()
    if not limpo:
        raise HTTPException(status_code=422, detail="O nome do departamento é obrigatório.")
    if len(limpo) > 80:
        raise HTTPException(status_code=422, detail="Nome muito longo (máximo 80 caracteres).")
    return limpo


# `require_admin`, e nao `require_admin_ou_lider`: a unica tela que consome
# isto hoje e /usuarios, que ja e de admin. Quando o lider precisar ver os
# proprios setores (etapa 4), a rota certa e outra, escopada a ele -- esta
# devolve TODOS os departamentos, e alcance total nao e o do lider.
@app.get("/api/departamentos", dependencies=[Depends(require_admin)])
def listar_departamentos() -> List[dict]:
    """
    Setores com os líderes e quantas pessoas têm.

    A contagem vem junto porque é o que responde "posso apagar este?" sem um
    segundo clique — e apagar um setor com gente dentro deixa essas pessoas
    sem departamento, o que ninguém quer descobrir depois.
    """
    from app.settings_state import _supabase

    sb = _supabase()
    if not sb:
        return []
    try:
        deps = sb.table("departamento").select("id, nome, criado_em").execute().data or []
        lids = sb.table("departamento_lider").select("departamento_id, user_id").execute().data or []
        pessoas = sb.table("users").select("id, full_name, email, departamento_id, ativo, role").execute().data or []
    except Exception:  # noqa: BLE001
        logger.exception("Falha ao listar departamentos")
        raise HTTPException(status_code=503, detail="Não foi possível listar os departamentos.")

    por_id = {str(u["id"]): u for u in pessoas}
    membros: Dict[str, int] = defaultdict(int)
    for u in pessoas:
        if u.get("departamento_id"):
            membros[str(u["departamento_id"])] += 1

    lideres: Dict[str, List[dict]] = defaultdict(list)
    for l in lids:
        u = por_id.get(str(l.get("user_id")))
        if not u:
            continue
        lideres[str(l.get("departamento_id"))].append({
            "id": str(u["id"]),
            "nome": u.get("full_name") or u.get("email"),
            "ativo": bool(conta_ativa(u)),
        })

    saida = []
    for d in deps:
        did = str(d["id"])
        saida.append({
            "id": did,
            "nome": d.get("nome"),
            "criado_em": d.get("criado_em"),
            "lideres": sorted(lideres.get(did, []), key=lambda x: x["nome"] or ""),
            "membros": membros.get(did, 0),
        })
    return sorted(saida, key=lambda x: (x["nome"] or "").lower())


@app.post("/api/departamentos", dependencies=[Depends(require_admin)])
def criar_departamento(body: DepartamentoBody) -> dict:
    from app.settings_state import _supabase

    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Sistema sem Supabase configurado.")
    nome = _nome_de_departamento(body.nome)
    try:
        r = sb.table("departamento").insert({"nome": nome}).execute()
    except Exception as e:  # noqa: BLE001
        # O índice único é sobre lower(btrim(nome)). A mensagem crua do
        # PostgREST diria "duplicate key value violates unique constraint", que
        # não ajuda quem está olhando um campo de texto.
        if "duplicate" in str(e).lower() or "unique" in str(e).lower():
            raise HTTPException(status_code=409, detail=f"Já existe um departamento chamado {nome}.")
        logger.exception("Falha ao criar departamento")
        raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True, "id": str((r.data or [{}])[0].get("id", ""))}


@app.put("/api/departamentos/{dep_id}", dependencies=[Depends(require_admin)])
def renomear_departamento(dep_id: str, body: DepartamentoBody) -> dict:
    from app.settings_state import _supabase

    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Sistema sem Supabase configurado.")
    nome = _nome_de_departamento(body.nome)
    try:
        sb.table("departamento").update({"nome": nome}).eq("id", dep_id).execute()
    except Exception as e:  # noqa: BLE001
        if "duplicate" in str(e).lower() or "unique" in str(e).lower():
            raise HTTPException(status_code=409, detail=f"Já existe um departamento chamado {nome}.")
        raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True}


@app.delete("/api/departamentos/{dep_id}", dependencies=[Depends(require_admin)])
def apagar_departamento(dep_id: str) -> dict:
    """
    Apaga o setor. As pessoas dele ficam SEM departamento, não são apagadas —
    é o `ON DELETE SET NULL` da migration, e a escolha é deliberada: perder o
    vínculo é corrigível na tela, perder as contas não.

    As lideranças caem junto (`ON DELETE CASCADE`): liderança de um setor que
    não existe mais daria alcance sobre nada e confundiria a leitura.
    """
    from app.settings_state import _supabase

    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Sistema sem Supabase configurado.")
    try:
        sb.table("departamento").delete().eq("id", dep_id).execute()
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True}


@app.put("/api/departamentos/{dep_id}/lideres", dependencies=[Depends(require_admin)])
def definir_lideres(dep_id: str, body: DepartamentoLideresBody) -> dict:
    """
    Substitui a lista de líderes do setor.

    Substitui em vez de somar porque a tela mostra a lista inteira: se o
    servidor só acrescentasse, tirar alguém exigiria uma rota a mais e a tela
    passaria a mentir sobre o que salvou.
    """
    from app.settings_state import _supabase

    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Sistema sem Supabase configurado.")

    ids = [str(x).strip() for x in (body.lideres or []) if str(x).strip()]
    if len(set(ids)) != len(ids):
        raise HTTPException(status_code=422, detail="A mesma pessoa aparece duas vezes na lista.")

    if ids:
        try:
            achados = sb.table("users").select("id, role, ativo").execute().data or []
        except Exception:  # noqa: BLE001
            raise HTTPException(status_code=503, detail="Não foi possível validar os líderes agora.")
        por_id = {str(u["id"]): u for u in achados}
        for uid in ids:
            u = por_id.get(uid)
            if not u:
                raise HTTPException(status_code=422, detail="Um dos líderes escolhidos não existe.")
            if not conta_ativa(u):
                # Líder desativado não entra no portal, então o setor ficaria
                # com um responsável que não consegue liberar nada — a mesma
                # situação de não ter líder, mas parecendo resolvida.
                raise HTTPException(
                    status_code=422,
                    detail="Não é possível designar uma conta desativada como líder.",
                )

    try:
        sb.table("departamento_lider").delete().eq("departamento_id", dep_id).execute()
        if ids:
            sb.table("departamento_lider").insert(
                [{"departamento_id": dep_id, "user_id": uid} for uid in ids]
            ).execute()
    except Exception as e:  # noqa: BLE001
        logger.exception("Falha ao definir líderes")
        raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True, "lideres": len(ids)}


@app.get("/api/health")
def health() -> dict:
    """
    Estado de configuração do ambiente.

    Devolve apenas BOOLEANOS de "está configurado?", nunca valores — a rota é
    pública. Serve para conferir um deploy (Vercel/Render) sem descobrir por
    tentativa e erro: o .env não sobe no deploy, então cada ambiente precisa
    ter suas variáveis definidas no painel da plataforma.
    """
    return {
        "ok": True,
        "supabase": supabase_configured(),
        "api_key_required": bool(config.API_KEY),
        # Sem esta chave o cofre não funciona: /upload-pfx falha no primeiro
        # certificado que o agente tentar enviar.
        "cert_vault_key_configurada": bool(
            config.CERT_ENCRYPTION_KEY and len(config.CERT_ENCRYPTION_KEY) == 64
        ),
        # Chave da SENHA do PFX. Sem ela — ou igual à do PFX, que a aplicação
        # recusa — `upsert_pfx` levanta e /upload-pfx devolve 500. O sintoma
        # aparece longe daqui: o cofre mantém o registro antigo, sem senha, e o
        # instalador avulso falha na máquina do usuário com "o portal não
        # enviou a senha". Sem estes dois campos, descobrir isso exige ler o
        # agent.log de um servidor.
        "cert_senha_key_configurada": bool(
            config.CERT_PASSWORD_ENCRYPTION_KEY
            and len(config.CERT_PASSWORD_ENCRYPTION_KEY) == 64
        ),
        "cert_senha_key_distinta": bool(
            config.CERT_PASSWORD_ENCRYPTION_KEY
            and config.CERT_PASSWORD_ENCRYPTION_KEY != config.CERT_ENCRYPTION_KEY
        ),
        # Ausente, a chave do SMTP é derivada da JWT_SECRET_KEY — o que faz a
        # senha SMTP parar de descriptografar se a JWT diferir entre ambientes.
        "smtp_key_dedicada": bool(os.getenv("ENCRYPTION_KEY")),
        "jwt_configurado": bool(os.getenv("JWT_SECRET_KEY")),
    }


def _settings_dict(s: PortalSettings) -> dict:
    return {
        "source_folder": s.source_folder,
        "expired_folder": s.expired_folder,
        "machine_id": s.machine_id,
        "effective_source": str(s.effective_source()),
        "effective_expired": str(s.effective_expired()),
        "supabase": supabase_configured(),
        "persistence": (
            "supabase+data/portal_settings.json"
            if supabase_configured()
            else "data/portal_settings.json"
        ),
        "smtp_host": s.smtp_host,
        "smtp_port": s.smtp_port,
        "smtp_user": s.smtp_user,
        "smtp_password_set": bool(s.smtp_password_encrypted),
        "smtp_use_tls": s.smtp_use_tls,
        "smtp_use_ssl": s.smtp_use_ssl,
        "instalador_nome_template": s.instalador_nome_template,
        # `efetivo` é o que realmente vale agora, com o padrão já resolvido —
        # a tela precisa mostrar isso, não o campo em branco.
        "instalador_nome_efetivo": s.instalador_nome_template or cert_installer.TEMPLATE_NOME_PADRAO,
        "install_token_ttl_min": s.install_token_ttl_min,
        "install_token_ttl_efetivo": cert_installer.ttl_do_token(),
        "trilha_retencao_dias": s.trilha_retencao_dias,
        "smtp_from_email": s.smtp_from_email,
        "smtp_alerts_enabled": s.smtp_alerts_enabled,
    }


@app.get("/api/settings", dependencies=[Depends(require_auth)])
def get_settings() -> dict:
    s = load_settings()
    return _settings_dict(s)


@app.put("/api/settings", dependencies=[Depends(require_admin)])
def put_settings(body: SettingsBody) -> dict:
    try:
        validate_smtp_config(body.smtp_use_tls, body.smtp_use_ssl)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    # Recusar na hora de salvar é a única defesa possível para o template: um
    # template sem {token} produz um .exe que abre e não instala nada, e o
    # sintoma aparece na máquina do usuário final.
    try:
        template = cert_installer.validar_template_nome(body.instalador_nome_template)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    ttl = int(body.install_token_ttl_min or 0)
    if ttl and not (cert_installer.TTL_TOKEN_MIN <= ttl <= cert_installer.TTL_TOKEN_MAX):
        raise HTTPException(
            status_code=422,
            detail=(
                f"Validade do token: use entre {cert_installer.TTL_TOKEN_MIN} e "
                f"{cert_installer.TTL_TOKEN_MAX} minutos, ou 0 para o padrão."
            ),
        )

    retencao = int(body.trilha_retencao_dias or 0)
    if retencao < 0:
        raise HTTPException(status_code=422, detail="Retenção não pode ser negativa.")

    
    old = load_settings()
    enc_password = old.smtp_password_encrypted
    if body.smtp_password is not None and body.smtp_password.strip() != "":
        try:
            enc_password = encrypt_password(body.smtp_password.strip())
        except Exception as e:
            raise HTTPException(status_code=500, detail="Erro ao criptografar senha SMTP")
            
    s = PortalSettings(
        source_folder=body.source_folder.strip(),
        expired_folder=body.expired_folder.strip(),
        machine_id=body.machine_id.strip() or "default",
        smtp_host=body.smtp_host.strip(),
        smtp_port=body.smtp_port,
        smtp_user=body.smtp_user.strip(),
        smtp_password_encrypted=enc_password,
        smtp_use_tls=body.smtp_use_tls,
        smtp_use_ssl=body.smtp_use_ssl,
        smtp_from_email=body.smtp_from_email.strip(),
        smtp_alerts_enabled=body.smtp_alerts_enabled,
        instalador_nome_template=template,
        install_token_ttl_min=ttl,
        trilha_retencao_dias=retencao,
    )
    save_settings(s)
    return _settings_dict(s)


# ══════════════════════════════════════════════════════════════════════════
# Recuperação de senha por código (A8 da auditoria de UI/UX)
# ══════════════════════════════════════════════════════════════════════════

# Resposta única para todos os desfechos do pedido: e-mail inexistente, conta
# desativada, envio bem-sucedido e até estouro do teto de pedidos. Distinguir
# transformaria a tela num detector de quem tem conta no sistema — e a lista de
# quem tem conta aqui é a lista de quem administra certificados de clientes.
RESPOSTA_GENERICA = (
    "Se houver uma conta com esse e-mail, enviamos um código de 6 dígitos. "
    "Ele vale por 15 minutos."
)

CODIGO_INVALIDO = "Código inválido ou expirado. Peça um novo se precisar."


class SenhaCodigoBody(BaseModel):
    email: str


class SenhaVerificarBody(BaseModel):
    email: str
    codigo: str


class SenhaRedefinirBody(BaseModel):
    email: str
    codigo: str
    password: str


def _conta_para_reset(sb, email: str) -> Optional[dict]:
    """
    A conta que pode receber código: existe e está ativa.

    Desativada não recebe — redefinir senha não pode ser caminho de volta para
    quem foi removido do portal. De fora não dá para distinguir, porque a
    resposta é a mesma.
    """
    try:
        r = (
            sb.table("users")
            .select("id, email, full_name, role, ativo")
            .eq("email", email)
            .limit(1)
            .execute()
        )
    except Exception:  # noqa: BLE001
        logger.exception("Falha ao procurar conta para recuperação de senha")
        return None
    linhas = r.data or []
    if not linhas:
        return None
    conta = linhas[0]
    return conta if auth.conta_ativa(conta) else None


def _enviar_codigo_por_email(conta: dict, codigo: str) -> None:
    """Manda o código. Levanta se o SMTP falhar — o chamador decide o que fazer."""
    s = load_settings()
    if not s.smtp_host:
        raise RuntimeError("SMTP não configurado")

    base = (os.getenv("PORTAL_BASE_URL") or "").strip().rstrip("/")
    # Link só de conveniência, e só se o endereço vier de configuração. Nunca
    # do cabeçalho `Host`: ele é controlado por quem chama, e um Host forjado
    # faria o portal mandar a própria vítima para o site do atacante.
    atalho = (
        f'<p style="margin:16px 0 0;">'
        f'<a href="{base}/login">Abrir o portal</a></p>' if base else ""
    )
    nome = html.escape(str(conta.get("full_name") or "").strip() or "Olá")

    smtp_service.send_smtp_email(
        host=s.smtp_host,
        port=s.smtp_port,
        user=s.smtp_user,
        password_enc=s.smtp_password_encrypted,
        use_tls=s.smtp_use_tls,
        use_ssl=s.smtp_use_ssl,
        from_email=s.smtp_from_email,
        to_email=str(conta["email"]),
        subject="Código para redefinir sua senha",
        html_content=(
            f"<p>{nome},</p>"
            "<p>Recebemos um pedido para redefinir a senha da sua conta no "
            "Monitor de Certificados. Use o código abaixo:</p>"
            f'<p style="font-size:28px;letter-spacing:6px;font-weight:700;'
            f'margin:24px 0;">{codigo}</p>'
            f"<p>Ele vale por {senha_reset.VALIDADE_MIN} minutos e só pode ser "
            "usado uma vez.</p>"
            "<p><strong>Se não foi você que pediu</strong>, ignore este e-mail: "
            "sua senha continua a mesma. Ninguém consegue trocá-la sem este "
            "código.</p>"
            f"{atalho}"
        ),
    )


@app.post("/api/senha/codigo")
def senha_pedir_codigo(body: SenhaCodigoBody, request: Request) -> dict:
    """
    Pede um código de redefinição. **Responde sempre a mesma coisa.**

    O 200 genérico é o ponto: qualquer variação de mensagem, status ou tempo
    de resposta entre "existe" e "não existe" vira enumeração de contas.
    """
    from app.settings_state import _supabase

    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Sistema sem Supabase configurado.")

    ip = request.client.host if request and request.client else None
    email = (body.email or "").strip().lower()
    if not email:
        return {"ok": True, "message": RESPOSTA_GENERICA}

    conta = _conta_para_reset(sb, email)
    if not conta:
        # Log em nível de info, sem alarde: e-mail digitado errado é o caso
        # comum, e não um incidente.
        logger.info("Pedido de código para e-mail sem conta ativa.")
        return {"ok": True, "message": RESPOSTA_GENERICA}

    try:
        codigo = senha_reset.criar_codigo(str(conta["id"]), client_ip=ip)
    except senha_reset.LimiteDePedidos:
        # Mesma resposta de propósito: dizer "você pediu demais" confirmaria
        # que a conta existe, que é justamente o que o genérico esconde.
        logger.warning(
            "Teto de %d pedidos/hora atingido para uma conta.",
            senha_reset.MAX_PEDIDOS_HORA,
        )
        return {"ok": True, "message": RESPOSTA_GENERICA}
    except Exception:  # noqa: BLE001
        logger.exception("Falha ao criar código de redefinição")
        raise HTTPException(
            status_code=503,
            detail="Não foi possível gerar o código agora. Tente de novo em instantes.",
        )

    try:
        _enviar_codigo_por_email(conta, codigo)
    except Exception:  # noqa: BLE001
        # ERROR e não warning: a pessoa está olhando para uma tela que diz que
        # o código foi enviado, e ele não foi. Sem este log ninguém descobre.
        logger.exception(
            "Código gerado mas NÃO enviado — a pessoa vai esperar um e-mail que não chega."
        )

    return {"ok": True, "message": RESPOSTA_GENERICA}


@app.post("/api/senha/verificar")
def senha_verificar_codigo(body: SenhaVerificarBody) -> dict:
    """
    Confere o código **sem consumi-lo**, para a tela avançar antes de a pessoa
    digitar a senha nova.

    Sem este passo, um código errado só apareceria depois de ela preencher a
    senha duas vezes — e já teria queimado uma das três tentativas à toa.
    """
    from app.settings_state import _supabase

    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Sistema sem Supabase configurado.")

    conta = _conta_para_reset(sb, (body.email or "").strip().lower())
    if not conta:
        # Sem conta, não há código. Recusa com a mesma mensagem de código
        # errado, para não distinguir os dois casos.
        raise HTTPException(status_code=400, detail=CODIGO_INVALIDO)

    ok, _ = senha_reset.conferir(str(conta["id"]), body.codigo or "")
    if not ok:
        raise HTTPException(status_code=400, detail=CODIGO_INVALIDO)
    return {"ok": True}


@app.post("/api/senha/redefinir")
def senha_redefinir(body: SenhaRedefinirBody, request: Request) -> dict:
    """
    Consome o código e grava a senha nova.

    Reconfere o código aqui em vez de confiar no `/verificar`: aquele passo é
    conveniência de tela, não credencial. Quem chamar esta rota direto tem de
    apresentar o código do mesmo jeito.
    """
    from app.settings_state import _supabase

    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Sistema sem Supabase configurado.")

    nova = (body.password or "").strip()
    if len(nova) < SENHA_MINIMA:
        raise HTTPException(
            status_code=422,
            detail=f"A senha precisa ter no mínimo {SENHA_MINIMA} caracteres.",
        )

    email = (body.email or "").strip().lower()
    conta = _conta_para_reset(sb, email)
    if not conta:
        raise HTTPException(status_code=400, detail=CODIGO_INVALIDO)

    if not senha_reset.consumir(str(conta["id"]), body.codigo or ""):
        raise HTTPException(status_code=400, detail=CODIGO_INVALIDO)

    agora = datetime.now(timezone.utc).isoformat()
    try:
        sb.table("users").update(
            {
                "password_hash": auth.get_password_hash(nova),
                # Carimbado na MESMA gravação da senha: separá-los abriria uma
                # janela em que a senha já mudou mas as sessões antigas ainda
                # valem — que é exatamente o que esta coluna existe para fechar.
                "senha_alterada_em": agora,
                # Aqui foi a própria pessoa quem escolheu, com um código que só
                # ela recebeu. Não há nada a cobrar depois.
                "deve_trocar_senha": False,
            }
        ).eq("id", conta["id"]).execute()
    except Exception as e:  # noqa: BLE001
        logger.exception("Falha ao gravar a senha nova")
        raise HTTPException(status_code=400, detail=str(e))

    atividade.registrar(
        atividade.EVENTO_SENHA_REDEFINIDA,
        user_id=str(conta["id"]),
        user_email=email,
        client_ip=request.client.host if request and request.client else None,
    )
    return {
        "ok": True,
        "message": "Senha alterada. Entre com a nova senha.",
    }


class SenhaTrocarBody(BaseModel):
    senha_atual: str
    nova_senha: str


@app.post("/api/senha/trocar")
def senha_trocar(
    body: SenhaTrocarBody,
    request: Request,
    token: auth.TokenData = Depends(require_auth),
) -> dict:
    """
    A própria pessoa troca a senha. É a única rota que responde com senha
    provisória — ver `ROTAS_COM_SENHA_PROVISORIA`.

    Exige a senha atual mesmo já estando autenticada. Parece redundante, e não
    é: com a sessão aberta e a máquina destravada, qualquer um que sente na
    cadeira definiria a senha nova sem saber a antiga, e a pessoa perderia a
    conta para quem passou por ali.
    """
    from app.settings_state import _supabase

    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Sistema sem Supabase configurado.")

    nova = (body.nova_senha or "").strip()
    if len(nova) < SENHA_MINIMA:
        raise HTTPException(
            status_code=422,
            detail=f"A senha precisa ter no mínimo {SENHA_MINIMA} caracteres.",
        )

    uid = _user_id_da_sessao(token)
    if not uid:
        raise HTTPException(status_code=401, detail=SESSAO_ENCERRADA)

    try:
        r = sb.table("users").select("password_hash").eq("id", uid).limit(1).execute()
    except Exception:  # noqa: BLE001
        logger.exception("Falha ao ler a conta para troca de senha")
        raise HTTPException(status_code=503, detail="Não foi possível trocar a senha agora.")

    linhas = r.data or []
    if not linhas or not auth.verify_password(body.senha_atual or "", linhas[0]["password_hash"]):
        raise HTTPException(status_code=400, detail="A senha atual não confere.")

    if auth.verify_password(nova, linhas[0]["password_hash"]):
        # Sem isto, "trocar a senha" seria satisfeito repetindo a provisória —
        # e a senha que outra pessoa conhece continuaria valendo, agora com a
        # flag desligada e ninguém mais cobrando a troca.
        raise HTTPException(
            status_code=422,
            detail="A senha nova precisa ser diferente da atual.",
        )

    agora = datetime.now(timezone.utc).isoformat()
    try:
        sb.table("users").update(
            {
                "password_hash": auth.get_password_hash(nova),
                "senha_alterada_em": agora,
                "deve_trocar_senha": False,
            }
        ).eq("id", uid).execute()
    except Exception as e:  # noqa: BLE001
        logger.exception("Falha ao gravar a senha nova")
        raise HTTPException(status_code=400, detail=str(e))

    atividade.registrar(
        atividade.EVENTO_SENHA_REDEFINIDA,
        user_id=uid,
        user_email=token.email or "",
        client_ip=request.client.host if request and request.client else None,
        contexto={"origem": "troca_propria"},
    )
    # `senha_alterada_em` acabou de ser carimbado, então o token que fez esta
    # chamada morre na requisição seguinte — de propósito, é a mesma regra que
    # derruba sessão aberta em qualquer troca de senha. O front reconhece o 401
    # e manda para o login.
    return {"ok": True, "message": "Senha alterada. Entre novamente com a nova senha."}


class SmtpTestBody(BaseModel):
    target_email: str


@app.post("/api/settings/smtp/test", dependencies=[Depends(require_admin)])
def test_smtp_config(body: SmtpTestBody) -> dict:
    s = load_settings()
    if not s.smtp_host:
        raise HTTPException(status_code=400, detail="Servidor SMTP não configurado.")
    try:
        # Qualificado pelo módulo: `send_smtp_email` nunca esteve na lista de
        # imports deste arquivo, então a rota levantava NameError em vez de
        # enviar. Só não aparecia porque ninguém clicava em "Enviar Teste".
        smtp_service.send_smtp_email(
            host=s.smtp_host,
            port=s.smtp_port,
            user=s.smtp_user,
            password_enc=s.smtp_password_encrypted,
            use_tls=s.smtp_use_tls,
            use_ssl=s.smtp_use_ssl,
            from_email=s.smtp_from_email,
            to_email=body.target_email.strip(),
            subject="Monitor de Certificados - E-mail de Teste",
            html_content="<p>Olá! Este é um e-mail de teste enviado a partir do seu <strong>Monitor de Certificados</strong> para validar as configurações de SMTP.</p>"
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True, "message": "E-mail de teste enviado com sucesso!"}


@app.post("/api/settings/alerts/trigger", dependencies=[Depends(require_admin)])
def trigger_alerts_manually() -> dict:
    try:
        stats = trigger_all_alerts()
        return {"ok": True, "stats": stats}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/cron/alerts")
def cron_alerts(request: Request) -> dict:
    """
    Disparo agendado dos alertas, chamado pelo Cron do Vercel.

    Existe porque o laço do `lifespan` não roda em serverless: cada requisição
    instancia a função e a encerra, então o `asyncio.create_task` morre antes
    dos 60s do primeiro disparo. No Render, com processo vivo, o laço bastava;
    no Vercel ninguém nunca chamaria `trigger_all_alerts`.

    Autenticação por CRON_SECRET em vez de JWT: o Cron do Vercel não faz login.
    Quando a variável CRON_SECRET existe no projeto, o Vercel envia
    `Authorization: Bearer <CRON_SECRET>` automaticamente em cada chamada.

    Falha FECHADA se a variável não estiver definida: uma rota que dispara
    envio de e-mail em massa não pode ficar aberta a quem descobrir a URL só
    porque alguém esqueceu de configurar o segredo.
    """
    segredo = (os.getenv("CRON_SECRET") or "").strip()
    if not segredo:
        logger.error("CRON_SECRET não configurada — disparo agendado recusado.")
        raise HTTPException(
            status_code=503,
            detail="CRON_SECRET não configurada no ambiente.",
        )

    enviado = (request.headers.get("authorization") or "").strip()
    esperado = f"Bearer {segredo}"
    # compare_digest evita vazar o segredo pelo tempo de resposta.
    if not secrets.compare_digest(enviado, esperado):
        logger.warning("Chamada ao cron de alertas com credencial inválida.")
        raise HTTPException(status_code=401, detail="Não autorizado.")

    try:
        stats = trigger_all_alerts()
        logger.info(f"Cron de alertas concluído: {stats}")
    except Exception as e:
        logger.exception("Falha no cron de alertas")
        raise HTTPException(status_code=500, detail=str(e))

    # Expurgo do install_log pendurado no mesmo disparo diário, e não num cron
    # próprio: os planos da Vercel limitam o número de crons, e um segundo
    # agendamento poderia simplesmente não ser criado — falha silenciosa numa
    # rotina de LGPD é o pior lugar para tê-la.
    #
    # Try/except próprio de propósito: apagar log é acessório, e não pode
    # derrubar o envio de alertas, que é o motivo de a rota existir. Se falhar,
    # aparece na resposta em vez de sumir.
    try:
        expurgo = {
            "install_log": cert_installer.expurgar_install_log(),
            "user_activity": atividade.expurgar(),
            # O cofre entra no mesmo ciclo: chave privada de certificado
            # vencido ou removido da pasta é passivo puro, e o acervo só
            # crescia porque nada a tirava.
            "cofre": cert_installer.expurgar_cofre(),
        }
    except Exception as e:  # noqa: BLE001
        logger.exception("Falha no expurgo da trilha")
        expurgo = {"executado": False, "motivo": str(e)}

    return {"ok": True, "stats": stats, "expurgo": expurgo}


@app.get("/api/colaborador/notificacoes", dependencies=[Depends(require_auth)])
def get_user_notifications(token: auth.TokenData = Depends(require_auth)) -> dict:
    try:
        # Devolve lista limitada + totais separados: antes eram 519 itens
        # (167 KB) a cada poll de 60s, com os acionáveis no fim da lista.
        return build_notifications_payload(
            token.email, token.role, _user_id_da_sessao(token)
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# Enfileirar comando para o agente e acao de operacao, e a unica chamadora e
# `configuracao.html` — pagina que so admin ve. Estava sob `require_auth`, que
# aceita QUALQUER autenticado: um operador comum podia mandar o servidor
# reescanear ou mover certificados. Ver docs/PLANO_niveis_de_acesso.md §1.
@app.post("/api/agent/commands", dependencies=[Depends(require_admin)])
def enqueue_agent_command(body: EnqueueCommandBody) -> dict:
    """
    Enfileira um comando para o agente Windows (poll em GET /api/agent/next).
    Comandos: mover_vencidos, rescan, ping. Use machine_id alinhado ao agente (ou * para qualquer um).
    """
    try:
        cid = enqueue(body.machine_id.strip() or "default", body.command.strip())
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e)) from e
    return {"ok": True, "id": cid, "command": body.command.strip()}


@app.get("/api/agent/next", dependencies=[Depends(require_agent_or_admin)])
def agent_next_command(
    machine_id: str = Query("default", description="ID da máquina do agente"),
) -> dict:
    """
    O agente chama isto no início de cada ciclo: retira um comando em fila ou null.
    """
    q = pop_next_for_agent(machine_id)
    if not q:
        return {"command": None, "id": None}
    out = {"command": q.command, "id": q.id, "machine_id": q.machine_id}
    # `payload` carrega o token de uso único de instalar_certificados. A linha
    # da fila já foi removida no pop, então o token só trafega uma vez, para o
    # agente autenticado que o solicitou.
    if q.payload:
        out["payload"] = q.payload
    return out


@app.get("/api/agent/queue", dependencies=[Depends(require_admin)])
def agent_queue_list() -> dict:
    """Lista comandos ainda pendentes (monitorização no portal)."""
    return {"pendentes": list_pending(), "comandos_validos": sorted(COMMANDS)}


@app.get("/api/certificados", dependencies=[Depends(require_auth)])
def listar_certificados(
    fonte: str = Query(
        "auto",
        description="auto | remoto | local",
    ),
    pagina: Optional[int] = Query(None, ge=1, description="Com por_pagina, ativa paginação no servidor"),
    por_pagina: Optional[int] = Query(None, ge=1, le=2000),
    filtro_status: str = Query("todos", description="todos | validos | prestes_vencer | vencidos | erros"),
    busca: Optional[str] = Query(None, max_length=400),
    todas_filtradas: bool = Query(False, description="Exportação: todos os itens do filtro (até LISTAGEM_EXPORT_MAX)"),
) -> JSONResponse:
    """
    * auto: usa o último snapshot ingerido se existir; senão leitura local.
    * remoto: só snapshot (404 se vazio).
    * local: sempre leitura no disco do servidor (pastas efetivas da config).

    Com ``pagina`` e ``por_pagina`` na query, devolve ``paginacao`` e ``resumo`` (painel).
    Sem esses parâmetros, mantém o comportamento antigo: lista completa em ``itens``.
    """
    try:
        sets = load_settings()
        snap = get_latest_snapshot()
        base = _list_certificados_payload(sets, snap, fonte)
        base["itens"] = ordenar_por_titular(base.get("itens") or [])

        paged = pagina is not None and por_pagina is not None
        if not paged and not todas_filtradas:
            return JSONResponse({**base, "supabase": supabase_configured()})

        now = datetime.now(timezone.utc)
        thirty = now + timedelta(days=30)
        enriched = [_enrich_cert_item_dashboard_flags(it, now, thirty) for it in (base.get("itens") or [])]
        filtered = [
            it
            for it in enriched
            if _dashboard_filtro_status_match(it, filtro_status) and _dashboard_busca_match(it, busca or "")
        ]
        resumo = _dashboard_resumo_counts(filtered)

        if todas_filtradas:
            lista_truncada = len(filtered) > LISTAGEM_EXPORT_MAX
            return JSONResponse(
                {
                    **base,
                    "itens": filtered[:LISTAGEM_EXPORT_MAX],
                    "resumo": resumo,
                    "lista_truncada": lista_truncada,
                    "supabase": supabase_configured(),
                }
            )

        pp = int(por_pagina or 20)
        total = len(filtered)
        total_pags = max(1, (total + pp - 1) // pp) if total else 1
        pagina_in = min(max(1, int(pagina or 1)), total_pags)
        off = (pagina_in - 1) * pp
        page_items = filtered[off : off + pp]
        return JSONResponse(
            {
                **base,
                "itens": page_items,
                "resumo": resumo,
                "paginacao": {
                    "pagina": pagina_in,
                    "total_paginas": total_pags,
                    "total_itens": total,
                    "por_pagina": pp,
                    # Permite ao portal avisar sobre truncamento ANTES de exportar,
                    # em vez de só depois que o arquivo já foi gerado.
                    "export_max": LISTAGEM_EXPORT_MAX,
                },
                "supabase": supabase_configured(),
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Erro em GET /api/certificados (fonte=%s)", fonte)
        raise HTTPException(
            status_code=500,
            detail="Falha ao listar certificados. Veja o terminal do uvicorn. Resumo: " + str(e),
        ) from e


def _escape_ilike_pattern(val: str) -> str:
    """Evita que % e _ do usuário interfiram com ILIKE."""
    return val.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


def _parse_iso_utc(iso_value: Optional[str]) -> datetime:
    if not iso_value:
        return datetime.min.replace(tzinfo=timezone.utc)
    s = str(iso_value).strip()
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except ValueError:
        return datetime.min.replace(tzinfo=timezone.utc)


def _normalize_ingest_items_status(items: List[dict], now: datetime) -> List[dict]:
    """
    Alinha gravacao com o criterio do painel: se not_after ja passou (UTC),
    marca status como expirado mesmo quando o agente enviou ok (ex.: relogio local desfasado).
    Nao altera erro, fora_do_padrao nem itens sem not_after valido.
    """
    min_dt = datetime.min.replace(tzinfo=timezone.utc)
    out: List[dict] = []
    for raw in items:
        it = dict(raw)
        s = str(it.get("status") or "").strip().lower()
        if s == CertStatus.OK.value:
            na = it.get("not_after")
            if na:
                exp = _parse_iso_utc(str(na))
                if exp > min_dt and exp < now:
                    it["status"] = CertStatus.EXPIRED.value
        out.append(it)
    return out


def _digits_only_doc(value: Any) -> str:
    return re.sub(r"\D", "", str(value or ""))


def _normalize_name_dup(value: Any) -> str:
    t = str(value or "").strip().lower()
    t = unicodedata.normalize("NFD", t)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return " ".join(t.split())


def _fingerprint_hex_from_row(row: dict) -> str:
    """SHA-256 (hex) do fingerprint do certificado; aceita chave antiga `cert_sha256` nos snapshots."""
    v = row.get("fingerprint_sha256") or row.get("cert_sha256")
    if v is None or not str(v).strip():
        return ""
    return str(v).strip().lower()


def _item_resumo_duplicidade(it: dict) -> dict[str, Any]:
    fp = it.get("fingerprint_sha256") or it.get("cert_sha256")
    return {
        "file_name": it.get("file_name"),
        "nome": it.get("nome") or it.get("display_name"),
        "documento": it.get("documento_formatado") or it.get("documento_numero"),
        "documento_numero": it.get("documento_numero"),
        "not_after": it.get("not_after"),
        "not_before": it.get("not_before"),
        "status": it.get("status"),
        "path": it.get("path"),
        "subject": it.get("subject"),
        "issuer": it.get("issuer"),
        "serial_number": it.get("serial_number"),
        "fingerprint_sha256": fp,
    }


def _fingerprint_hex_resumo(m: dict) -> str:
    v = m.get("fingerprint_sha256")
    if v is None or not str(v).strip():
        return ""
    return str(v).strip().lower()


def _filtrar_grupo_documento_apos_fingerprint(members: List[dict]) -> List[dict]:
    """
    Remove da lista «mesmo documento» os arquivos que já entram no agrupamento
    por fingerprint (2+ com o mesmo SHA-256). A duplicidade criptográfica é a
    validação definitiva; o grupo por documento fica para CPF/CNPJ igual sem
    fingerprint ou com certificados distintos (ex.: renovação).
    """
    by_fp: dict[str, List[dict]] = defaultdict(list)
    sem_fp: List[dict] = []
    for m in members:
        fp = _fingerprint_hex_resumo(m)
        if not fp:
            sem_fp.append(m)
        else:
            by_fp[fp].append(m)
    kept: List[dict] = []
    kept.extend(sem_fp)
    for _fp, grupo in by_fp.items():
        if len(grupo) < 2:
            kept.extend(grupo)
    return kept


def _agrupar_duplicidades(
    rows: List[dict],
) -> Tuple[List[dict], List[dict], List[dict]]:
    """
    Deteta duplicidades no mesmo inventário (último snapshot ou scan local):
    - mesmo CNPJ/CPF (11+ dígitos) em mais de um arquivo (exceto quando a duplicidade
      já é explicada só por fingerprint — aí fica só em certificados idênticos);
    - certificados idênticos: mesmo fingerprint (SHA-256 do DER) em mais de um arquivo;
    - nomes muito semelhantes (SequenceMatcher) só quando não existe fingerprint
      no inventário (export antigo do agente ou leitura falhou).
    """
    by_doc: dict[str, List[dict]] = defaultdict(list)
    for it in rows:
        d = _digits_only_doc(it.get("documento_numero") or it.get("documento_formatado"))
        if len(d) >= 11:
            by_doc[d].append(_item_resumo_duplicidade(it))

    grupos_documento: List[dict] = []
    for doc_digits, members in by_doc.items():
        if len(members) < 2:
            continue
        filtrados = _filtrar_grupo_documento_apos_fingerprint(members)
        if len(filtrados) < 2:
            continue
        exib = next((m.get("documento") for m in filtrados if m.get("documento")), doc_digits)
        grupos_documento.append(
            {
                "tipo": "documento",
                "documento_digitos": doc_digits,
                "documento_exibicao": exib,
                "itens": filtrados,
            }
        )

    by_fp: dict[str, List[dict]] = defaultdict(list)
    for it in rows:
        fp = _fingerprint_hex_from_row(it)
        if not fp:
            continue
        by_fp[fp].append(_item_resumo_duplicidade(it))

    grupos_cert_igual: List[dict] = []
    for fp_hex, members in by_fp.items():
        if len(members) < 2:
            continue
        grupos_cert_igual.append(
            {
                "tipo": "certificado_igual",
                "fingerprint_sha256": fp_hex,
                "itens": members,
            }
        )

    n = len(rows)
    parent = list(range(n))

    def find(a: int) -> int:
        while parent[a] != a:
            parent[a] = parent[parent[a]]
            a = parent[a]
        return a

    def union(a: int, b: int) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for i in range(n):
        for j in range(i + 1, n):
            if _fingerprint_hex_from_row(rows[i]) or _fingerprint_hex_from_row(rows[j]):
                continue
            fi = str(rows[i].get("file_name") or "").strip().lower()
            fj = str(rows[j].get("file_name") or "").strip().lower()
            if not fi or fi == fj:
                continue
            di = _digits_only_doc(
                rows[i].get("documento_numero") or rows[i].get("documento_formatado")
            )
            dj = _digits_only_doc(
                rows[j].get("documento_numero") or rows[j].get("documento_formatado")
            )
            if len(di) >= 11 and len(dj) >= 11 and di == dj:
                continue
            ni = _normalize_name_dup(
                rows[i].get("nome") or rows[i].get("display_name") or rows[i].get("file_name")
            )
            nj = _normalize_name_dup(
                rows[j].get("nome") or rows[j].get("display_name") or rows[j].get("file_name")
            )
            if len(ni) < 5 or len(nj) < 5:
                continue
            if SequenceMatcher(None, ni, nj).ratio() >= 0.86:
                union(i, j)

    roots: dict[int, List[int]] = defaultdict(list)
    for i in range(n):
        roots[find(i)].append(i)

    grupos_nome: List[dict] = []
    for _root, idxs in roots.items():
        if len(idxs) < 2:
            continue
        members = [_item_resumo_duplicidade(rows[k]) for k in idxs]
        nomes_cur = [
            _normalize_name_dup(
                rows[k].get("nome") or rows[k].get("display_name") or rows[k].get("file_name")
            )
            for k in idxs
        ]
        rotulo = max(nomes_cur, key=len) if nomes_cur else "—"
        grupos_nome.append({"tipo": "nome_similar", "rotulo": rotulo[:120], "itens": members})

    return grupos_documento, grupos_nome, grupos_cert_igual


def _doc_norm(v: Any) -> str:
    return re.sub(r"\D+", "", str(v or ""))


def _parse_dt_or_min(v: Any) -> datetime:
    return _parse_iso_utc(str(v or ""))


def _status_prioridade(status: str) -> int:
    s = str(status or "").lower()
    if s in ("ok", "valido", "válido"):
        return 3
    if s in ("expirado", "vencido"):
        return 2
    if s in ("erro",):
        return 1
    return 0


def _lista_base_docs_historico() -> List[dict]:
    # Limitamos a 100 snapshots para não sobrecarregar a memória no Render free tier.
    # Para colaboradores, apenas os snapshots recentes são relevantes.
    hist = historico_certificados(limite_snapshots=100)
    rows = hist.get("itens", [])
    grupos: Dict[str, dict] = {}
    for it in rows:
        doc = _doc_norm(it.get("documento"))
        if not doc:
            continue
        atual = grupos.get(doc)
        cand = {
            "documento": it.get("documento") or doc,
            "documento_digitos": doc,
            "nome": it.get("nome") or "—",
            "status_ultimo": str(it.get("status_ultimo") or "").lower(),
            "vencimento_certificado": it.get("vencimento_certificado"),
            "ultima_data_registrada": it.get("ultima_data_registrada"),
        }
        if not atual:
            grupos[doc] = cand
            continue
        pa = _status_prioridade(atual.get("status_ultimo", ""))
        pc = _status_prioridade(cand.get("status_ultimo", ""))
        if pc > pa:
            grupos[doc] = cand
            continue
        if pc == pa:
            da = _parse_dt_or_min(atual.get("ultima_data_registrada"))
            dc = _parse_dt_or_min(cand.get("ultima_data_registrada"))
            if dc > da:
                grupos[doc] = cand
    return sorted(grupos.values(), key=lambda x: (x.get("nome") or "").lower())


def _painel_docs_selecionados(doc_ids: List[str]) -> List[dict]:
    base = _lista_base_docs_historico()
    by_doc = {str(it.get("documento_digitos")): it for it in base}
    now = datetime.now(timezone.utc)
    out: List[dict] = []
    for d in doc_ids:
        it = by_doc.get(d)
        if not it:
            out.append(
                {
                    "documento_digitos": d,
                    "documento": d,
                    "nome": "Não encontrado no inventário atual",
                    "status": "nao_encontrado",
                    "vencimento_certificado": None,
                    "dias_restantes": None,
                }
            )
            continue
        v_iso = it.get("vencimento_certificado")
        v_dt = _parse_iso_utc(v_iso) if v_iso else datetime.min.replace(tzinfo=timezone.utc)
        dias = (v_dt.date() - now.date()).days if v_iso else None
        
        status_ult = str(it.get("status_ultimo") or "").lower()
        if status_ult == "expirado" or (v_iso and v_dt < now):
            status = "vencido"
        elif status_ult == "erro":
            status = "erro"
        elif status_ult == "fora_do_padrao":
            status = "fora_do_padrao"
        elif v_iso and dias is not None and 0 <= dias <= 30:
            status = "expirando"
        else:
            status = "ativo"

        out.append(
            {
                "documento_digitos": d,
                "documento": it.get("documento") or d,
                "nome": it.get("nome") or "—",
                "status": status,
                "vencimento_certificado": v_iso,
                "dias_restantes": dias,
            }
        )
    out.sort(
        key=lambda x: (
            0 if x.get("status") == "vencido" else 1,
            x.get("dias_restantes") if x.get("dias_restantes") is not None else 10**9,
        )
    )
    return out


class ColaboradorSelecaoBody(BaseModel):
    documentos: List[str] = Field(default_factory=list)


@app.get("/api/colaborador/certificados/opcoes", dependencies=[Depends(require_auth)])
def colaborador_opcoes_certificados(_token: auth.TokenData = Depends(require_auth)) -> dict:
    itens = _lista_base_docs_historico()
    now = datetime.now(timezone.utc)
    out = []
    for it in itens:
        v_iso = it.get("vencimento_certificado")
        v_dt = _parse_iso_utc(v_iso) if v_iso else datetime.min.replace(tzinfo=timezone.utc)
        dias = (v_dt.date() - now.date()).days if v_iso else None
        
        status_ult = str(it.get("status_ultimo") or "").lower()
        if status_ult == "expirado" or (v_iso and v_dt < now):
            status = "vencido"
        elif status_ult == "erro":
            status = "erro"
        elif status_ult == "fora_do_padrao":
            status = "fora_do_padrao"
        elif v_iso and dias is not None and 0 <= dias <= 30:
            status = "expirando"
        else:
            status = "ativo"
            
        out.append({
            **it,
            "status": status
        })
    return {"itens": out, "total": len(out)}


@app.get("/api/colaborador/certificados/selecionados", dependencies=[Depends(require_auth)])
def colaborador_get_selecionados(token: auth.TokenData = Depends(require_auth)) -> dict:
    email = (token.email or "").strip().lower()
    docs = load_colaborador_selecao(email, _user_id_da_sessao(token))
    return {"documentos": docs, "total": len(docs)}


@app.put("/api/colaborador/certificados/selecionados", dependencies=[Depends(require_auth)])
def colaborador_put_selecionados(
    body: ColaboradorSelecaoBody, token: auth.TokenData = Depends(require_auth)
) -> dict:
    email = (token.email or "").strip().lower()
    docs = sorted({_doc_norm(x) for x in body.documentos if _doc_norm(x)})
    save_colaborador_selecao(email, docs, _user_id_da_sessao(token))
    return {"ok": True, "documentos": docs, "total": len(docs)}


@app.get("/api/colaborador/certificados/painel", dependencies=[Depends(require_auth)])
def colaborador_painel_certificados(token: auth.TokenData = Depends(require_auth)) -> dict:
    email = (token.email or "").strip().lower()
    docs = load_colaborador_selecao(email, _user_id_da_sessao(token))
    itens = _painel_docs_selecionados(docs)
    return {"itens": itens, "total": len(itens)}


@app.get("/api/certificados/duplicidades", dependencies=[Depends(require_auth)])
def certificados_duplicidades() -> dict[str, Any]:
    """
    Analisa o último snapshot recebido (dados atuais do agente) ou, na ausência,
    o scan local no servidor, e devolve grupos de possíveis duplicados.
    """
    snap = get_latest_snapshot()
    origem = "ultimo_snapshot"
    scanned_at: Optional[str] = None
    if snap and (snap.get("items") or []):
        raw_items: List[dict] = list(snap.get("items") or [])
        scanned_at = str(snap.get("scanned_at") or "") or None
    else:
        sets = load_settings()
        raw_items = [cert_to_public_dict(c) for c in scan_folder(sets.effective_source())]
        origem = "scan_local_servidor"
        scanned_at = datetime.now(timezone.utc).isoformat()

    rows = [it for it in raw_items if str(it.get("file_name") or "").strip()]
    gd, gn, gci = _agrupar_duplicidades(rows)
    return {
        "origem_dados": origem,
        "scanned_at": scanned_at,
        "total_itens_analisados": len(rows),
        "grupos_documento": gd,
        "grupos_nome_similar": gn,
        "grupos_certificado_igual": gci,
        "total_grupos_documento": len(gd),
        "total_grupos_nome_similar": len(gn),
        "total_grupos_certificado_igual": len(gci),
    }


def _historico_merge_snapshot_into_agregados(snap: dict[str, Any], agregados: Dict[str, dict]) -> None:
    """Acumula itens de um snapshot no mapa por file_name (mantém linha do scan mais recente)."""
    scanned_at = snap.get("scanned_at") or datetime.now(timezone.utc).isoformat()
    scanned_dt = _parse_iso_utc(scanned_at)
    for it in (snap.get("items") or []):
        file_name = str(it.get("file_name") or "").strip()
        if not file_name:
            continue
        key = file_name.lower()
        atual = agregados.get(key)
        if (not atual) or (scanned_dt > atual["_dt"]):
            doc_raw = (
                it.get("documento_formatado") or it.get("documento_numero") or it.get("documento")
            )
            agregados[key] = {
                "_dt": scanned_dt,
                "file_name": file_name,
                "nome": it.get("nome") or it.get("display_name") or file_name,
                "status_ultimo": it.get("status"),
                "documento": doc_raw,
                "vencimento_certificado": it.get("not_after"),
                "ultima_data_registrada": scanned_dt.isoformat(),
            }


def _historico_carregar_agregados(limite_snapshots: int) -> Tuple[Dict[str, dict], int]:
    """
    Percorre snapshots (Supabase em lotes ou arquivo local) e devolve agregação por file_name.
    Resultado pode vir de cache em RAM (TTL configurável) por (Supabase ativo, limite).
    """
    from app.settings_state import _supabase

    sb = _supabase()
    uses_sb = sb is not None

    def _build() -> Tuple[Dict[str, dict], int]:
        agregados: Dict[str, dict] = {}
        snapshots_lidos = 0
        if sb:
            try:
                page_size = 50
                offset = 0
                while snapshots_lidos < limite_snapshots:
                    chunk = min(page_size, limite_snapshots - snapshots_lidos)
                    end = offset + chunk - 1
                    r = (
                        sb.table("cert_snapshots")
                        .select("scanned_at, items")
                        .order("scanned_at", desc=True)
                        .range(offset, end)
                        .execute()
                    )
                    rows = r.data or []
                    if not rows:
                        break
                    for snap in rows:
                        _historico_merge_snapshot_into_agregados(snap, agregados)
                    snapshots_lidos += len(rows)
                    offset += len(rows)
                    if len(rows) < chunk:
                        break
            except Exception as e:  # noqa: BLE001
                logger.exception("Falha ao ler histórico no Supabase")
                raise HTTPException(status_code=500, detail=f"Falha ao ler histórico: {e}") from e
        else:
            snap = get_latest_snapshot()
            if snap:
                _historico_merge_snapshot_into_agregados(snap, agregados)
                snapshots_lidos = 1
        return agregados, snapshots_lidos

    return _historico_cache_get_or_build(uses_sb, limite_snapshots, _build)


def _historico_itens_visualizacao(agregados: Dict[str, dict]) -> List[dict]:
    linhas = sorted(agregados.values(), key=lambda x: x["_dt"], reverse=True)
    return [{k: v for k, v in row.items() if k != "_dt"} for row in linhas]


def _historico_filtrar_busca(itens: List[dict], busca_raw: str) -> List[dict]:
    q = str(busca_raw or "").strip().lower()
    if not q:
        return itens
    out: List[dict] = []
    for it in itens:
        haystack = (
            f"{it.get('file_name') or ''} {it.get('nome') or ''} {it.get('documento') or ''} "
            f"{it.get('ultima_data_registrada') or ''} {it.get('vencimento_certificado') or ''}"
        ).lower()
        if q in haystack:
            out.append(it)
    return out


def _vencidos_filtrar_busca(rows: List[dict], busca_raw: str) -> List[dict]:
    if not str(busca_raw or "").strip():
        return rows
    raw = str(busca_raw).strip()
    bt = _painel_busca_normalizada(raw)
    bd = re.sub(r"\D", "", raw)
    out: List[dict] = []
    for it in rows:
        nome = _painel_busca_normalizada(it.get("nome") or "")
        doc_txt = _painel_busca_normalizada(it.get("documento") or "")
        doc_d = _digits_only_doc(it.get("documento") or "")
        venc_txt = _painel_busca_normalizada(it.get("vencimento_certificado") or "")
        vn = _digits_only_doc(str(it.get("vencimento_certificado") or ""))
        if bt and bt in nome:
            out.append(it)
            continue
        if bt and bt in doc_txt:
            out.append(it)
            continue
        if bt and bt in venc_txt:
            out.append(it)
            continue
        if bd and bd in doc_d:
            out.append(it)
            continue
        if bd and bd in vn:
            out.append(it)
            continue
    return out


def _cert_history_fetch_all(sb: Any) -> List[dict[str, Any]]:
    """Lê todas as linhas de cert_history (PostgREST limita ~1000 por pedido sem range)."""
    cols = (
        "file_name, nome, documento, status_ultimo, "
        "vencimento_certificado, ultima_data_registrada"
    )
    batch = 1000
    out: List[dict[str, Any]] = []
    offset = 0
    while True:
        r = (
            sb.table("cert_history")
            .select(cols)
            .order("ultima_data_registrada", desc=True)
            .range(offset, offset + batch - 1)
            .execute()
        )
        chunk = r.data or []
        out.extend(chunk)
        if len(chunk) < batch:
            break
        offset += batch
    return out


def historico_certificados(
    limite_snapshots: int = 500,
    *,
    offset: Optional[int] = None,
    limit: Optional[int] = None,
    busca: Optional[str] = None,
) -> dict:
    """
    Lista certificados já mapeados em algum momento, com a última data registrada.
    Com ``limit`` definido, lê só uma página de ``cert_history`` (menos carga).

    Chamadas internas devem omitir ``limit`` para obter a lista completa.
    """
    from app.settings_state import _supabase

    pagination = limit is not None
    offset = max(0, int(offset or 0))
    page_limit = int(limit) if pagination else None
    busca_txt = str(busca).strip() if busca else ""

    def _normalize_rows(rows_raw: List[dict[str, Any]]) -> List[dict[str, Any]]:
        return [
            {
                "file_name":              row.get("file_name"),
                "nome":                   row.get("nome"),
                "documento":              row.get("documento"),
                "status_ultimo":          row.get("status_ultimo"),
                "vencimento_certificado": row.get("vencimento_certificado"),
                "ultima_data_registrada": row.get("ultima_data_registrada"),
            }
            for row in rows_raw
        ]

    def _apply_cert_history_or_busca(qb: Any) -> Any:
        if not busca_txt:
            return qb
        pat = "%" + _escape_ilike_pattern(busca_txt) + "%"
        filt = f"nome.ilike.{pat},file_name.ilike.{pat},documento.ilike.{pat}"
        return qb.or_(filt)

    sb = _supabase()
    if sb:
        _use_history_table = True
        rows_non_paginated: List[dict[str, Any]] = []
        try:
            if pagination and page_limit is not None:
                qc = _apply_cert_history_or_busca(
                    sb.table("cert_history").select("file_name", count="exact", head=True)
                )
                c_r = qc.execute()
                total_count = c_r.count if c_r.count is not None else 0

                if total_count > 0 or busca_txt:
                    qp = _apply_cert_history_or_busca(
                        sb.table("cert_history")
                        .select(
                            "file_name, nome, documento, status_ultimo, "
                            "vencimento_certificado, ultima_data_registrada",
                        )
                        .order("ultima_data_registrada", desc=True)
                    )
                    hi = offset + page_limit - 1
                    p_r = qp.range(offset, hi).execute()
                    return {
                        "itens": _normalize_rows(p_r.data or []),
                        "total": total_count,
                        "offset": offset,
                        "limit": page_limit,
                        "snapshots_lidos": 0,
                        "fonte": "cert_history",
                    }

                # total_count == 0 e sem texto de busca → tentar snapshots (dados legados)
            else:
                rows_non_paginated = _cert_history_fetch_all(sb)
        except Exception as e:  # noqa: BLE001
            err_str = str(e)
            if "PGRST205" in err_str or "cert_history" in err_str:
                logger.warning(
                    "Tabela cert_history não encontrada; usando fallback de snapshots. "
                    "Execute supabase/migrations/20260504_cert_history.sql no Supabase."
                )
                _use_history_table = False
                rows_non_paginated = []
            else:
                logger.exception("Falha inesperada ao ler cert_history no Supabase")
                raise HTTPException(status_code=500, detail=f"Falha ao ler histórico: {e}") from e

        if _use_history_table and rows_non_paginated and not pagination:
            itens = _normalize_rows(rows_non_paginated)
            return {
                "itens": itens,
                "total": len(itens),
                "offset": 0,
                "limit": len(itens),
                "snapshots_lidos": 0,
                "fonte": "cert_history",
            }

    agregados, snapshots_lidos = _historico_carregar_agregados(limite_snapshots)
    itens = _historico_itens_visualizacao(agregados)
    itens = _historico_filtrar_busca(itens, busca_txt)

    if pagination and page_limit is not None:
        total_f = len(itens)
        fatia = itens[offset : offset + page_limit]
        return {
            "itens": fatia,
            "total": total_f,
            "offset": offset,
            "limit": page_limit,
            "snapshots_lidos": snapshots_lidos,
            "fonte": "snapshots",
        }

    return {
        "itens": itens,
        "total": len(itens),
        "offset": 0,
        "limit": len(itens),
        "snapshots_lidos": snapshots_lidos,
        "fonte": "snapshots",
    }


@app.get("/api/certificados/historico", dependencies=[Depends(require_auth)])
def historico_certificados_http(
    limite_snapshots: int = Query(
        500,
        ge=1,
        le=2000,
        description="Máximo de snapshots no fallback (quando cert_history está vazio)",
    ),
    pagina: int = Query(1, ge=1, description="Página (1-based)"),
    por_pagina: int = Query(20, ge=1, le=2000, description="Registros por página"),
    todas_filtradas: bool = Query(
        False,
        description="Quando true, devolve toda a lista filtrada (exportação; pode truncar)",
    ),
    busca: Optional[str] = Query(
        None,
        max_length=200,
        description="Filtro parcial em nome, arquivo ou documento",
    ),
) -> dict:
    b = busca.strip() if busca else None
    if todas_filtradas:
        raw = historico_certificados(limite_snapshots, offset=None, limit=None, busca=b)
        itens = list(raw.get("itens") or [])
        lista_truncada = len(itens) > LISTAGEM_EXPORT_MAX
        return {
            "itens": itens[:LISTAGEM_EXPORT_MAX],
            "total": len(itens),
            "snapshots_lidos": raw.get("snapshots_lidos", 0),
            "fonte": raw.get("fonte"),
            "lista_truncada": lista_truncada,
        }

    off = (pagina - 1) * por_pagina
    raw = historico_certificados(
        limite_snapshots,
        offset=off,
        limit=por_pagina,
        busca=b,
    )
    total = int(raw.get("total") or 0)
    total_pags = max(1, (total + por_pagina - 1) // por_pagina) if total else 1
    pagina_out = min(max(1, pagina), total_pags)
    out = dict(raw)
    out["paginacao"] = {
        "pagina": pagina_out,
        "total_paginas": total_pags,
        "total_itens": total,
        "por_pagina": por_pagina,
        "export_max": LISTAGEM_EXPORT_MAX,
    }
    return out


@app.get("/api/certificados/vencidos", dependencies=[Depends(require_auth)])
def vencidos_certificados(
    data_inicio: Optional[str] = Query(None, description="Data inicial (YYYY-MM-DD) pelo vencimento"),
    data_fim: Optional[str] = Query(None, description="Data final (YYYY-MM-DD) pelo vencimento"),
    pagina: int = Query(1, ge=1),
    por_pagina: int = Query(20, ge=1, le=2000),
    todas_filtradas: bool = Query(False, description="Lista completa filtrada (exportação; pode truncar)"),
    busca: Optional[str] = Query(None, max_length=200),
    limite_snapshots: int = Query(500, ge=1, le=2000, description="Quantidade máxima de snapshots lidos"),
) -> dict:
    # Vencidos precisa de uma agregação tão ampla quanto a do histórico (fallback snapshots).
    lim_hist = max(limite_snapshots, config.HISTORICO_LIMITE_SNAPSHOTS)
    hist = historico_certificados(lim_hist, offset=None, limit=None, busca=None)
    itens_hist = hist.get("itens", [])
    inicio_dt = _parse_iso_utc(data_inicio + "T00:00:00+00:00") if data_inicio else None
    fim_dt = _parse_iso_utc(data_fim + "T23:59:59+00:00") if data_fim else None
    busca_txt = str(busca or "").strip() or None

    now_utc = datetime.now(timezone.utc)
    min_dt = datetime.min.replace(tzinfo=timezone.utc)

    def _conta_como_certificado_vencido(it: dict) -> bool:
        """Inclui `expirado`/`vencido` no status ou data de validade já passada (alinha ao painel)."""
        s = str(it.get("status_ultimo") or "").lower()
        if s in ("expirado", "vencido"):
            return True
        venc_dt = _parse_iso_utc(it.get("vencimento_certificado"))
        if venc_dt <= min_dt:
            return False
        return venc_dt < now_utc

    venc_filtrados: List[dict] = []
    for it in itens_hist:
        if not _conta_como_certificado_vencido(it):
            continue
        venc_dt = _parse_iso_utc(it.get("vencimento_certificado"))
        s_low = str(it.get("status_ultimo") or "").lower()
        if inicio_dt or fim_dt:
            if venc_dt <= min_dt:
                if s_low not in ("expirado", "vencido"):
                    continue
            else:
                if inicio_dt and venc_dt < inicio_dt:
                    continue
                if fim_dt and venc_dt > fim_dt:
                    continue
        venc_filtrados.append(it)

    venc_filtrados = _vencidos_filtrar_busca(venc_filtrados, busca_txt or "")

    anos_cnt: defaultdict[int, int] = defaultdict(int)
    for it in venc_filtrados:
        venc_dt = _parse_iso_utc(it.get("vencimento_certificado"))
        y = int(venc_dt.year)
        if y < 1900:
            continue
        anos_cnt[y] += 1
    resumo_anos = [{"ano": ano, "total": anos_cnt[ano]} for ano in sorted(anos_cnt.keys(), reverse=True)]

    total = len(venc_filtrados)
    if todas_filtradas:
        lista_truncada = total > LISTAGEM_EXPORT_MAX
        return {
            "itens": venc_filtrados[:LISTAGEM_EXPORT_MAX],
            "total": total,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "snapshots_lidos": hist.get("snapshots_lidos", 0),
            "lista_truncada": lista_truncada,
            "resumo_anos": resumo_anos,
        }

    total_pags = max(1, (total + por_pagina - 1) // por_pagina) if total else 1
    pagina_out = min(max(1, pagina), total_pags)
    off_pg = (pagina_out - 1) * por_pagina
    pagina_slice = venc_filtrados[off_pg : off_pg + por_pagina]

    return {
        "itens": pagina_slice,
        "total": total,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "snapshots_lidos": hist.get("snapshots_lidos", 0),
        "resumo_anos": resumo_anos,
        "paginacao": {
            "pagina": pagina_out,
            "total_paginas": total_pags,
            "total_itens": total,
            "por_pagina": por_pagina,
            "export_max": LISTAGEM_EXPORT_MAX,
        },
    }


# Ingestao de inventario: quem alimenta e o agente (`agent/run_agent.py`), e
# `require_auth` aceitava qualquer autenticado — um operador comum podia
# sobrescrever o inventario inteiro. `require_agent_or_admin` e a mesma guarda
# que `upload-pfx`, `redeem` e `report` ja usam, e o agente ja passa por ela em
# producao (o cofre tem 491 certificados que so chegaram por `upload-pfx`).
@app.post("/api/ingest", dependencies=[Depends(require_agent_or_admin)])
def ingest(body: IngestBody, background_tasks: BackgroundTasks) -> dict:
    """
    Recebe o resultado de um scan feito no Windows (agente em segundo plano).
    Persiste no Supabase (ou em data/last_ingest.json se o Supabase não estiver configurado).
    Também faz upsert na tabela materializada cert_history para acelerar o histórico.
    """
    machine_id = body.machine_id.strip() or "default"
    scanned = datetime.now(timezone.utc)
    items = _normalize_ingest_items_status(body.items, scanned)
    save_snapshot(
        machine_id=machine_id,
        source_folder=body.source_folder.strip(),
        expired_folder=body.expired_folder.strip(),
        items=items,
    )
    # Atualiza a tabela materializada — operação rápida, não bloqueia o retorno
    upsert_cert_history(
        machine_id=machine_id,
        scanned_iso=scanned.isoformat(),
        items=items,
    )
    # Dispara e-mails de alerta em segundo plano para não bloquear a resposta do agente
    background_tasks.add_task(trigger_all_alerts)
    return {
        "ok": True,
        "itens_recebidos": len(body.items),
        "grava_em": "supabase" if supabase_configured() else "arquivo local (data/last_ingest.json)",
    }


# Move arquivo de certificado no sistema de arquivos do servidor. Estava sob
# `require_auth` — qualquer autenticado. Nenhum template ou script do portal
# chama esta rota; ela e acionada fora da UI, e quem a aciona e operacao.
@app.post("/api/mover-vencidos", dependencies=[Depends(require_admin)])
def mover_vencidos() -> JSONResponse:
    """
    Só move arquivos no **mesmo** sistema de arquivos que corre o API (servidor acessa as pastas).
    Se a interface mostrar dados "remotos" vindos do agente, use o agendador no Windows
    (agente com --mover) para mover aí o disco local.
    """
    sets = load_settings()
    src = sets.effective_source()
    exp = sets.effective_expired()
    itens: List[CertInfo] = scan_folder(src)
    movidos: List[dict] = []
    erros: List[dict] = []

    for c in itens:
        if c.status != CertStatus.EXPIRED:
            continue
        try:
            novo = move_to_expired(c, exp)
            movidos.append({"de": str(c.path), "para": str(novo)})
        except OSError as e:
            erros.append({"arquivo": c.file_name, "erro": str(e)})

    return JSONResponse(
        {
            "movidos": movidos,
            "erros": erros,
            "total_movidos": len(movidos),
        }
    )


# =========================================================================
# LGPD / PRIVACY BY DESIGN - DIREITOS DOS TITULARES
# =========================================================================

@app.get("/api/users/me/export")
def export_my_data(token: auth.TokenData = Depends(require_auth)) -> dict:
    """[LGPD] Portabilidade e Consulta de Dados (Art. 18, incisos II e X)"""
    from app.settings_state import load_colaborador_selecao
    email = token.email
    docs = load_colaborador_selecao(email, _user_id_da_sessao(token))
    
    return {
        "titular": email,
        "vinculo_role": token.role,
        "documentos_monitorados_cnpj_cpf": docs,
        "exportado_em": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "aviso_lgpd": "Este arquivo contém seus dados pessoais conforme registro no sistema."
    }

@app.delete("/api/users/me/delete")
def delete_my_data(token: auth.TokenData = Depends(require_auth)) -> dict:
    """[LGPD] Direito ao Esquecimento / Eliminação (Art. 18, inciso VI)"""
    from app.settings_state import _supabase
    email = token.email
    sb = _supabase()
    
    # 1. Apagar seleções (Ações do usuário no app)
    if sb:
        # Pela identidade. O apagamento por `user_email` que acompanhava este
        # daqui saiu na fase 3c: a coluna deixou de ser escrita e some na 3d,
        # então continuar filtrando por ela seria apagar por um valor que o
        # portal já não grava — e daria a impressão de cobertura que não há.
        #
        # Sem `user_id` não dá para apagar nada com segurança, e numa rota de
        # LGPD isso não pode passar calado.
        uid = _user_id_da_sessao(token)
        if uid:
            sb.table("colaborador_cert_selecoes").delete().eq("user_id", uid).execute()
        else:
            logger.error(
                "Pedido de eliminação de %s não removeu as seleções: sessão sem "
                "user_id. O dado continua no banco.",
                email,
            )
    else:
        # Modo arquivo local: limpa a entrada
        from app.settings_state import save_colaborador_selecao
        save_colaborador_selecao(email, [])
        
    return {
        "status": "ok", 
        "message": "Seus dados operacionais associados foram permanentemente removidos."
    }


# ══════════════════════════════════════════════════════════════════════════
# MÓDULO INSTALADOR DE CERTIFICADOS DIGITAIS
# ══════════════════════════════════════════════════════════════════════════

from app import cert_installer


class UploadPfxRequest(BaseModel):
    """Payload enviado pelo agente com o PFX cifrado em trânsito."""
    fingerprint: str
    machine_id: str = "default"
    pfx_b64: str  # PFX em base64 (cifrado em trânsito via TLS)
    password: Optional[str] = None
    nome_titular: Optional[str] = None
    documento: Optional[str] = None
    documento_tipo: Optional[str] = None
    subject: Optional[str] = None
    not_before: Optional[str] = None
    not_after: Optional[str] = None
    friendly_name: Optional[str] = None


@app.post("/api/cert-installer/upload-pfx")
def upload_pfx(
    body: UploadPfxRequest,
    request: Request,
    token: auth.TokenData = Depends(require_agent_or_admin),
):
    """
    Recebe um PFX do agente, cifra com a chave do servidor e armazena.
    Chamado pelo agente durante o ciclo de scan.
    """
    import base64 as b64mod
    try:
        pfx_bytes = b64mod.b64decode(body.pfx_b64)
    except Exception:
        raise HTTPException(status_code=400, detail="pfx_b64 inválido")

    # Um PFX real tem poucos KB. 50 MB era um vetor de DoS via base64 no banco.
    if len(pfx_bytes) > 1 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="PFX excede 1 MB")

    # Barreira de servidor da custódia: mesmo que um agente desatualizado (ou
    # adulterado) envie o que não devia, só entra no cofre o que a política
    # permite. O filtro por machine_id é parte da barreira, não detalhe de
    # consulta — a custódia é por estação desde a chave composta.
    #
    # Sob opt-in, esta barreira e a lista que o agente consome eram a MESMA
    # consulta, e uma falha nela devolvia lista vazia: nada passava. Sob
    # opt-out isso se inverte, e é aqui que a tradução literal do código antigo
    # abriria o portão — "não consegui ler os bloqueios" viraria "nada está
    # bloqueado, pode gravar". Por isso `CustodiaIndisponivel` é tratada como
    # recusa explícita, e não cai no `except Exception` genérico lá embaixo.
    try:
        autorizados = cert_installer.fingerprints_autorizados(body.machine_id)
    except cert_installer.CustodiaIndisponivel as e:
        logger.warning("Upload recusado por custódia indisponível (%s): %s", body.machine_id, e)
        raise HTTPException(
            status_code=503,
            detail="Custódia indisponível; o envio será retentado no próximo ciclo.",
        )

    if body.fingerprint not in autorizados:
        raise HTTPException(
            status_code=403,
            detail=(
                "Certificado fora da custódia desta estação: está bloqueado, "
                "vencido, ilegível, ou ausente do último inventário."
            ),
        )

    try:
        cert_id = cert_installer.upsert_pfx(
            fingerprint=body.fingerprint,
            pfx_bytes=pfx_bytes,
            machine_id=body.machine_id,
            password=body.password,
            nome_titular=body.nome_titular,
            documento=body.documento,
            documento_tipo=body.documento_tipo,
            subject=body.subject,
            not_before=body.not_before,
            not_after=body.not_after,
            friendly_name=body.friendly_name,
        )
        return {"status": "ok", "id": cert_id, "fingerprint": body.fingerprint}
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception:
        logger.exception("Erro ao processar upload de PFX")
        raise HTTPException(status_code=500, detail="Erro interno ao armazenar PFX")


class VaultOptinRequest(BaseModel):
    """Admin autoriza um certificado a ter o PFX guardado no cofre."""
    fingerprint: str
    machine_id: str = "default"
    nome_titular: Optional[str] = None
    documento: Optional[str] = None


@app.get("/api/cert-installer/vault-optin")
def listar_vault_optin(
    machine_id: Optional[str] = Query(None),
    _token: auth.TokenData = Depends(require_agent_or_admin),
):
    """
    Fingerprints que esta máquina pode enviar ao cofre.

    **O caminho e o formato da resposta não mudaram na inversão de 15/08**, e
    isso é deliberado: o agente instalado no ANALISESRV é um `.exe` compilado
    que sempre perguntou "o que posso mandar?" e só agiu sobre a resposta.
    Trocar o que entra na lista — de "autorizados um a um" para "inventário
    válido menos bloqueados" — inverteu a custódia sem recompilar nada.

    Sem `machine_id` não há pergunta a responder: a custódia é por estação
    desde a chave composta, e uma lista global não significa nada aqui.
    """
    if not machine_id:
        raise HTTPException(
            status_code=422,
            detail="machine_id é obrigatório: a custódia é definida por estação.",
        )
    try:
        return {"fingerprints": sorted(cert_installer.fingerprints_autorizados(machine_id))}
    except cert_installer.CustodiaIndisponivel as e:
        # 503, nunca lista vazia. O agente trata não-200 como "não enviar
        # nada"; uma lista vazia ele trataria como "nada a enviar", que é o
        # mesmo efeito hoje — mas as duas respostas dizem coisas diferentes, e
        # confundi-las é como o opt-out vira falha aberta na próxima mudança.
        logger.warning("Custódia indisponível para %s: %s", machine_id, e)
        raise HTTPException(
            status_code=503,
            detail="Não foi possível determinar a custódia do cofre agora.",
        )


@app.post("/api/cert-installer/vault-optin", dependencies=[Depends(require_admin)])
def reativar_vault_custodia(
    body: VaultOptinRequest,
    _token: auth.TokenData = Depends(require_admin),
):
    """
    Devolve o certificado à custódia: apaga o bloqueio.

    Sob opt-in isto se chamava "autorizar" e gravava uma permissão. Sob opt-out
    a permissão é o padrão, então a ação equivalente é remover a exceção. O
    caminho continua o mesmo para não quebrar a tela; o que ele faz, não.

    O material volta sozinho na varredura seguinte — o cofre é derivado dos
    arquivos do ANALISESRV, não há o que restaurar aqui.
    """
    try:
        cert_installer.reativar_custodia(
            fingerprint=body.fingerprint,
            machine_id=body.machine_id,
        )
        return {"status": "ok", "fingerprint": body.fingerprint, "custodia": "ativa"}
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception:
        logger.exception("Erro ao reativar custódia do certificado")
        raise HTTPException(status_code=500, detail="Erro interno ao reativar custódia")


@app.delete("/api/cert-installer/vault-optin/{fingerprint}", dependencies=[Depends(require_admin)])
def bloquear_vault_custodia(
    fingerprint: str,
    machine_id: str = Query(..., min_length=1),
    motivo: Optional[str] = Query(None, max_length=300),
    token: auth.TokenData = Depends(require_admin),
):
    """
    Desativa a custódia: registra o bloqueio e APAGA o PFX — de UMA estação.

    **O bloqueio é o que faz isto durar.** Sob opt-in bastava apagar
    autorização e material: sem autorização o agente não reenviava. Sob
    opt-out, apagar só o material seria um botão que se desfaz sozinho — o
    certificado segue no inventário, volta a ser autorizado no ciclo seguinte
    e o PFX sobe de novo em até 24h.

    O `machine_id` é obrigatório: desde a chave composta `(machine_id,
    fingerprint)`, o mesmo certificado pode estar no cofre de várias estações,
    e a rota não tem como adivinhar qual delas o admin quer desativar. Sem o
    parâmetro a chamada é recusada, em vez de alcançar todas.
    """
    try:
        cert_installer.bloquear_custodia(
            fingerprint=fingerprint,
            machine_id=machine_id,
            bloqueado_por=token.email or "desconhecido",
            motivo=motivo,
        )
        return {
            "status": "ok",
            "fingerprint": fingerprint,
            "machine_id": machine_id,
            "pfx_removido": True,
            "custodia": "bloqueada",
        }
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception:
        logger.exception("Erro ao revogar certificado do cofre")
        raise HTTPException(status_code=500, detail="Erro interno ao revogar")


# Teto de certificados por token. Cada PFX ocupa ~10 KB no cofre e vai inteiro
# no bundle do instalador; sem limite, `certificate_ids` é um array livre e o
# download vira imprevisível. Recusar com número claro é melhor que entregar um
# arquivo de dezenas de MB que talvez nem baixe.
MAX_CERTIFICADOS_POR_TOKEN = 50


def _validar_pedido_de_instalacao(
    user_id: str, role: str, certificate_ids: List[str]
) -> None:
    """
    Tudo o que precisa valer antes de emitir um token de instalação.

    Função única de propósito: as duas rotas emissoras chamam exatamente esta,
    e `tests/test_carteira.py` lê o código para falhar se alguma rota nova
    emitir token sem passar por aqui. Duplicar a checagem seria o caminho
    natural, e a cópia que divergisse para o lado permissivo não daria sintoma
    nenhum — só entregaria certificado a quem não devia.

    `CarteiraIndisponivel` vira 503 e não 403: negar por falha de banco é o
    resultado seguro, mas dizer "você não tem permissão" a quem tem manda a
    pessoa procurar o gestor em vez de esperar o banco voltar.
    """
    if len(certificate_ids) > MAX_CERTIFICADOS_POR_TOKEN:
        raise HTTPException(
            status_code=422,
            detail=(
                f"Selecione no máximo {MAX_CERTIFICADOS_POR_TOKEN} certificados por "
                f"instalador ({len(certificate_ids)} selecionados)."
            ),
        )
    try:
        cert_installer.assegurar_carteira(user_id, role, certificate_ids)
    except cert_installer.ForaDaCarteira as e:
        logger.warning("Instalação negada por carteira (user=%s): %s", user_id, e)
        raise HTTPException(status_code=403, detail=str(e))
    except cert_installer.CarteiraIndisponivel as e:
        logger.warning("Carteira indisponível (user=%s): %s", user_id, e)
        raise HTTPException(
            status_code=503,
            detail="Não foi possível verificar suas permissões agora. Tente novamente.",
        )


ERRO_SEM_ALCANCE = (
    "Você não lidera nenhum departamento, então não há para quem liberar "
    "certificados. Peça a um administrador para incluí-lo como líder."
)


async def require_admin_ou_lider(
    token: auth.TokenData = Depends(require_auth),
) -> auth.TokenData:
    """
    Quem pode montar carteira: admin, ou quem lidera ao menos um departamento.

    Mudou em 18/08/2026. Antes bastava o papel `gestor`, e o alcance era total
    — qualquer gestor liberava qualquer cliente para qualquer operador. Agora o
    papel abre a porta e a LIDERANÇA define até onde se vai; cada rota confere
    o alvo com `cert_installer.pode_gerir`.

    Gestor sem liderança nenhuma é recusado aqui mesmo, com uma mensagem que
    diz o que fazer. Deixá-lo entrar numa tela onde toda ação falha depois
    seria pior: o sintoma viraria "não consigo salvar nada".
    """
    papel = (token.role or "").strip().lower()
    if papel in cert_installer.PAPEIS_COM_ALCANCE_TOTAL:
        return token
    if papel != "gestor":
        raise HTTPException(status_code=403, detail="Acesso restrito a gestores e administradores.")

    uid = _user_id_da_sessao(token)
    try:
        if uid and cert_installer.departamentos_que_lidera(uid):
            return token
    except cert_installer.AlcanceIndisponivel:
        # 503, e não 403: "não consegui verificar" não é "você não pode". Um
        # 403 aqui faria o líder acreditar que perdeu a permissão.
        raise HTTPException(
            status_code=503,
            detail="Não foi possível verificar seus departamentos. Tente de novo.",
        )
    raise HTTPException(status_code=403, detail=ERRO_SEM_ALCANCE)


def _exigir_alcance(token: auth.TokenData, alvo_id: str) -> None:
    """
    Barreira por PESSOA. `require_admin_ou_lider` só diz que o ator pode montar
    carteiras; esta diz de quem.

    Sem ela, um líder do Fiscal montaria a carteira de alguém do Contábil
    apenas trocando o `user_id` na chamada — a tela filtra, mas a tela não é a
    barreira.
    """
    try:
        if cert_installer.pode_gerir(_user_id_da_sessao(token) or "", token.role or "", alvo_id):
            return
    except cert_installer.AlcanceIndisponivel:
        raise HTTPException(
            status_code=503,
            detail="Não foi possível verificar seu alcance. Tente de novo.",
        )
    raise HTTPException(
        status_code=403,
        detail="Esta pessoa não está em um departamento que você lidera.",
    )


class CarteiraRequest(BaseModel):
    user_id: str
    documentos: List[str]


@app.get("/api/carteira/operadores")
def listar_operadores(
    token: auth.TokenData = Depends(require_admin_ou_lider),
) -> dict:
    """
    Quem pode receber carteira, com quantos documentos cada um já tem.

    Rota própria em vez de `/api/users`: aquela é de admin e devolve a linha
    inteira do usuário. O gestor precisa montar carteira **sem** poder
    administrar contas, e não tem por que ver o hash de senha de ninguém.
    """
    from app.settings_state import _supabase

    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Supabase não configurado")
    try:
        us = sb.table("users").select(
            "id, email, full_name, role, ativo, gestor_id, departamento_id"
        ).execute().data or []
        cart = sb.table("carteira").select("user_id").execute().data or []
    except Exception:
        logger.exception("Falha ao listar operadores")
        raise HTTPException(status_code=503, detail="Não foi possível listar os operadores.")

    from collections import Counter

    # O líder vê apenas quem ele alcança. Mostrar a lista inteira e recusar
    # depois seria a pior combinação: ele monta a carteira de alguém do outro
    # setor, clica em salvar, e só aí descobre — sem entender o critério.
    #
    # Filtrado aqui e conferido de novo em cada rota que age: esta é a tela, e
    # tela não é barreira. Quem chamar a API direto com outro `user_id` esbarra
    # em `_exigir_alcance`.
    eu = _user_id_da_sessao(token) or ""
    papel = (token.role or "").strip().lower()
    if papel in cert_installer.PAPEIS_COM_ALCANCE_TOTAL:
        visiveis = us
    else:
        try:
            meus = cert_installer.departamentos_que_lidera(eu)
        except cert_installer.AlcanceIndisponivel:
            raise HTTPException(
                status_code=503,
                detail="Não foi possível verificar seus departamentos. Tente de novo.",
            )
        visiveis = [
            u for u in us
            if str(u.get("id")) == eu
            or (u.get("departamento_id") and str(u["departamento_id"]) in meus)
        ]

    quantos = Counter(str(c.get("user_id")) for c in cart)
    return {
        "operadores": [
            {
                "id": str(u.get("id")),
                "email": u.get("email"),
                "full_name": u.get("full_name"),
                "role": u.get("role"),
                "ativo": auth.conta_ativa(u),
                "gestor_id": u.get("gestor_id"),
                "departamento_id": u.get("departamento_id"),
                "documentos": quantos.get(str(u.get("id")), 0),
            }
            for u in visiveis
        ]
    }


@app.get("/api/carteira/documentos", dependencies=[Depends(require_admin_ou_lider)])
def listar_documentos_atribuiveis(
    q: Optional[str] = Query(None, max_length=120),
    limite: int = Query(500, ge=1, le=2000),
) -> dict:
    """
    Universo de documentos para atribuir, filtrável por nome ou número.

    O teto era 500 e havia 491 clientes — a um cadastro de distância de
    truncar em silêncio, que é como a curva de vencimento perdeu 29
    certificados em 15/08. Subiu para 2000; a tela pede a lista inteira de
    uma vez (medido: 491 documentos = 33 KB em ~375 ms) e monta os dois
    painéis no cliente, então filtrar no servidor virou opcional.

    `total` continua vindo separado de `documentos` justamente para a tela
    poder dizer quando a lista foi cortada, em vez de parecer completa.
    """
    todos = cert_installer.universo_de_documentos()
    termo = (q or "").strip().lower()
    if termo:
        digitos = cert_installer.so_digitos(termo)
        todos = [
            d for d in todos
            if termo in (d["nome"] or "").lower()
            or (digitos and digitos in d["documento"])
        ]
    return {"total": len(todos), "documentos": todos[:limite]}


@app.get("/api/carteira/{user_id}")
def obter_carteira(
    user_id: str,
    token: auth.TokenData = Depends(require_admin_ou_lider),
) -> dict:
    """
    Documentos que este operador pode instalar, com a trilha de atribuição.

    A dependência saiu do decorador e virou parâmetro porque agora o token é
    USADO: `_exigir_alcance` precisa saber quem está perguntando. A trilha diz
    quem liberou o quê — informação de dentro do setor.
    """
    _exigir_alcance(token, user_id)
    try:
        linhas = cert_installer.detalhar_carteira(user_id)
    except cert_installer.CarteiraIndisponivel as e:
        raise HTTPException(status_code=503, detail=str(e))

    nomes = {d["documento"]: d["nome"] for d in cert_installer.universo_de_documentos()}
    return {
        "user_id": user_id,
        "documentos": sorted(l["documento"] for l in linhas),
        "itens": [
            {
                "documento": l["documento"],
                # Documento atribuído que não está mais no inventário continua
                # na carteira: a atribuição é uma decisão, e sumir com ela
                # esconderia que a pessoa tem acesso a algo que voltou depois.
                "nome": nomes.get(l["documento"], ""),
                "no_inventario": l["documento"] in nomes,
                "atribuido_por": l.get("atribuido_por_email"),
                "atribuido_em": l.get("atribuido_em"),
            }
            for l in linhas
        ],
    }


@app.post("/api/carteira")
def atribuir_carteira(
    body: CarteiraRequest,
    token: auth.TokenData = Depends(require_admin_ou_lider),
) -> dict:
    """
    Acrescenta documentos à carteira de um operador.

    Quem atribui fica registrado — e-mail inclusive, não só o UUID. Com o
    gestor podendo atribuir qualquer cliente do acervo, essa trilha é a única
    forma de reconstruir o que houve se uma conta de gestor for comprometida;
    guardar só o UUID a perderia no dia em que a conta fosse apagada.
    """
    _exigir_alcance(token, body.user_id)
    try:
        gravados = cert_installer.atribuir_carteira(
            user_id=body.user_id,
            documentos=body.documentos,
            atribuido_por=_user_id_da_sessao(token),
            atribuido_por_email=token.email or "desconhecido",
        )
        return {"status": "ok", "gravados": gravados}
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception:
        logger.exception("Erro ao atribuir carteira")
        raise HTTPException(status_code=500, detail="Erro interno ao atribuir carteira")


def _linhas_da_planilha(nome: str, raw: bytes) -> List[Dict[str, str]]:
    """
    Lê .csv ou .xlsx e devolve as linhas como dicionários de cabeçalho→valor.

    O .xlsx entra porque é o que sai do Excel sem passo extra — pedir "salve
    como CSV" antes de cada importação é exatamente o trabalho manual que
    esta rota existe para tirar. O `openpyxl` só é importado aqui: quem nunca
    importa planilha não paga o custo no cold start da Vercel.
    """
    if nome.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError:  # pragma: no cover - depende do ambiente de deploy
            raise HTTPException(
                status_code=503,
                detail="Leitura de .xlsx indisponível no servidor. Envie o arquivo como .csv.",
            )
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            linhas = list(ws.iter_rows(values_only=True))
        except Exception:
            raise HTTPException(status_code=422, detail="Não consegui ler a planilha .xlsx.")
        finally:
            try:
                wb.close()
            except Exception:
                pass
        if not linhas:
            raise HTTPException(status_code=422, detail="Planilha vazia.")
        cabecalho = [str(c or "").strip() for c in linhas[0]]
        return [
            {cabecalho[i]: ("" if v is None else str(v).strip())
             for i, v in enumerate(linha) if i < len(cabecalho)}
            for linha in linhas[1:]
        ]

    # CSV: mesmo tratamento do import de usuários — BOM do Excel e separador
    # `;` do pt-BR são a regra, não a exceção.
    texto = raw.decode("utf-8-sig", errors="replace")
    try:
        delim = csv.Sniffer().sniff(texto[:2048], delimiters=",;").delimiter
    except csv.Error:
        delim = ";"
    leitor = csv.DictReader(io.StringIO(texto), delimiter=delim)
    if not leitor.fieldnames:
        raise HTTPException(status_code=422, detail="CSV sem cabeçalho.")
    return [{(k or ""): (v or "") for k, v in linha.items()} for linha in leitor]


@app.post("/api/carteira/importar")
async def importar_carteiras(
    file: UploadFile = File(...),
    token: auth.TokenData = Depends(require_admin_ou_lider),
) -> dict:
    """
    Atribui carteiras em massa a partir de uma planilha de e-mail + documento.

    Montar carteira clicando cliente a cliente não escala: quem recebe uma
    lista pronta da operação copiava CNPJ por CNPJ na mão.

    O ALCANCE É CONFERIDO POR PESSOA, e não uma vez para o arquivo: o gestor
    importa só para quem ele lidera, exatamente como na tela. Uma linha fora
    do alcance vira erro DAQUELA linha — recusar o arquivo inteiro por causa
    de uma pessoa faria o gestor perder o trabalho já correto.

    Nada é gravado até o arquivo inteiro ser lido e validado: um arquivo com
    metade das linhas erradas deixaria a carteira em estado parcial, e o
    operador não teria como saber o que entrou.
    """
    nome = (file.filename or "").lower()
    if not (nome.endswith(".csv") or nome.endswith(".xlsx")):
        raise HTTPException(
            status_code=422,
            detail="Formato inválido. Envie a planilha em .xlsx ou .csv.",
        )

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=422, detail="Arquivo vazio.")
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Arquivo muito grande (limite de 5MB).")
    # Executável disfarçado. O `PK` do zip é legítimo aqui — todo .xlsx começa
    # com ele —, então a checagem é por extensão declarada.
    if raw.startswith(b"MZ") or raw.startswith(b"\x7fELF") or raw.startswith(b"%PDF"):
        raise HTTPException(status_code=422, detail="Conteúdo do arquivo suspeito.")
    if nome.endswith(".xlsx") and not raw.startswith(b"PK"):
        raise HTTPException(status_code=422, detail="Isto não é um .xlsx válido.")

    linhas = _linhas_da_planilha(nome, raw)
    if not linhas:
        raise HTTPException(status_code=422, detail="A planilha não tem nenhuma linha de dados.")

    mapa = {_norm_header(h): h for linha in linhas[:1] for h in linha}

    def coluna(*aliases: str) -> Optional[str]:
        for a in aliases:
            k = mapa.get(_norm_header(a))
            if k:
                return k
        return None

    col_email = coluna("email", "e-mail", "colaborador", "email do colaborador", "usuario")
    col_doc = coluna("documento", "cnpj", "cpf", "cnpj/cpf", "cnpj da empresa", "cliente")
    if not col_email or not col_doc:
        raise HTTPException(
            status_code=422,
            detail="A planilha precisa de duas colunas: e-mail do colaborador e CNPJ/CPF do cliente.",
        )

    from app.settings_state import _supabase

    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Supabase não configurado.")
    try:
        contas = sb.table("users").select("id, email").execute().data or []
    except Exception:
        logger.exception("Falha ao ler usuários na importação de carteiras")
        raise HTTPException(status_code=503, detail="Não foi possível ler os usuários.")
    por_email = {str(u.get("email") or "").strip().lower(): str(u.get("id")) for u in contas}

    universo = {d["documento"] for d in cert_installer.universo_de_documentos()}

    erros: List[Dict[str, Any]] = []
    por_usuario: Dict[str, set] = {}
    alcance: Dict[str, bool] = {}
    ator = _user_id_da_sessao(token) or ""

    for i, linha in enumerate(linhas, start=2):  # 1 é o cabeçalho
        email = (linha.get(col_email) or "").strip().lower()
        doc = cert_installer.so_digitos(linha.get(col_doc) or "")
        if not email and not doc:
            continue  # linha em branco no fim da planilha
        if not email or not doc:
            erros.append({"linha": i, "motivo": "Falta o e-mail ou o CNPJ/CPF."})
            continue

        user_id = por_email.get(email)
        if not user_id:
            erros.append({"linha": i, "motivo": f"Não existe usuário com o e-mail {email}."})
            continue

        if user_id not in alcance:
            try:
                alcance[user_id] = cert_installer.pode_gerir(ator, token.role or "", user_id)
            except cert_installer.AlcanceIndisponivel:
                raise HTTPException(
                    status_code=503,
                    detail="Não foi possível verificar seu alcance. Tente de novo.",
                )
        if not alcance[user_id]:
            erros.append({"linha": i, "motivo": f"{email} não está em um departamento que você lidera."})
            continue

        if doc not in universo:
            erros.append({"linha": i, "motivo": f"O documento {doc} não está no inventário."})
            continue

        por_usuario.setdefault(user_id, set()).add(doc)

    atribuidos = 0
    pessoas = 0
    for user_id, docs in por_usuario.items():
        try:
            atribuidos += cert_installer.atribuir_carteira(
                user_id=user_id,
                documentos=sorted(docs),
                atribuido_por=ator,
                atribuido_por_email=token.email or "desconhecido",
            )
            pessoas += 1
        except Exception:
            logger.exception("Falha ao gravar carteira importada")
            erros.append({"linha": 0, "motivo": "Falha ao gravar a carteira de um dos operadores."})

    return {
        "status": "ok",
        "atribuidos": atribuidos,
        "pessoas": pessoas,
        "linhas_lidas": len(linhas),
        "erros": erros,
    }


@app.delete("/api/carteira/{user_id}/{documento}")
def remover_carteira(
    user_id: str,
    documento: str,
    token: auth.TokenData = Depends(require_admin_ou_lider),
) -> dict:
    _exigir_alcance(token, user_id)
    try:
        cert_installer.remover_da_carteira(user_id, documento)
        return {"status": "ok", "user_id": user_id, "documento": documento}
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception:
        logger.exception("Erro ao remover da carteira")
        raise HTTPException(status_code=500, detail="Erro interno ao remover da carteira")


class ConfigInstaladorBody(BaseModel):
    instalador_nome_template: str = ""
    install_token_ttl_min: int = 0
    trilha_retencao_dias: int = 0


@app.put("/api/cert-installer/configuracao", dependencies=[Depends(require_admin)])
def salvar_config_instalador(body: ConfigInstaladorBody) -> dict:
    """
    Grava **só** as três configurações do módulo instalador.

    Rota própria em vez de reaproveitar `PUT /api/settings`: aquele monta um
    `PortalSettings` inteiro a partir do corpo, então uma tela que mandasse
    apenas estes três campos apagaria host, usuário e senha do SMTP — sem erro
    nenhum, e ninguém notaria até o próximo alerta não sair.

    Aqui a configuração atual é lida, três campos mudam, e o resto vai de volta
    como estava.
    """
    try:
        template = cert_installer.validar_template_nome(body.instalador_nome_template)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    ttl = int(body.install_token_ttl_min or 0)
    if ttl and not (cert_installer.TTL_TOKEN_MIN <= ttl <= cert_installer.TTL_TOKEN_MAX):
        raise HTTPException(
            status_code=422,
            detail=(
                f"Validade do token: use entre {cert_installer.TTL_TOKEN_MIN} e "
                f"{cert_installer.TTL_TOKEN_MAX} minutos, ou 0 para o padrão."
            ),
        )

    retencao = int(body.trilha_retencao_dias or 0)
    if retencao < 0:
        raise HTTPException(status_code=422, detail="Retenção não pode ser negativa.")

    atual = load_settings()
    atual.instalador_nome_template = template
    atual.install_token_ttl_min = ttl
    atual.trilha_retencao_dias = retencao
    save_settings(atual)
    return _settings_dict(atual)


@app.post("/api/cert-installer/expurgar-log", dependencies=[Depends(require_admin)])
def expurgar_log_agora() -> dict:
    """
    Roda o expurgo sob demanda, sem esperar o cron.

    Quem acabou de configurar a retenção precisa ver o efeito para confiar
    nela — e uma rotina de LGPD que só roda amanhã de manhã não dá para
    verificar antes de responder por ela.
    """
    return {
        "install_log": cert_installer.expurgar_install_log(),
        "user_activity": atividade.expurgar(),
        "cofre": cert_installer.expurgar_cofre(),
    }


@app.get("/api/dashboard", dependencies=[Depends(require_modulo("dashboard"))])
def dashboard_visao_geral(dias: int = Query(30, ge=1, le=365)) -> dict:
    """
    Os painéis baratos do dashboard, numa chamada (~1s).

    Separado das renovações de propósito: aquele precisa de dois snapshots
    completos (~1 MB) e os outros seis somam poucas dezenas de KB. Fazer o
    barato esperar o caro atrasaria toda a tela pelo painel menos urgente.
    """
    from app import dashboard

    return dashboard.visao_geral(dias)


@app.get("/api/dashboard/renovacoes", dependencies=[Depends(require_modulo("dashboard"))])
def dashboard_renovacoes(
    dias: int = Query(30, ge=1, le=365),
    machine_id: str = Query("ANALISESRV", min_length=1),
) -> dict:
    """
    Renovações: o inventário de hoje contra o de N dias atrás.

    Sai de `cert_snapshots`, não de `cert_history` — aquela é
    `upsert(on_conflict="file_name")` e guarda só o estado atual, então o valor
    anterior já foi sobrescrito e a conta daria zero.

    A resposta traz `referencia`: as varreduras têm lacunas, e pedir 30 dias
    pode devolver a comparação com uma de 54 dias atrás. Apresentar isso como
    "últimos 30 dias" seria mentira.
    """
    from app import dashboard

    return dashboard.painel_renovacoes(dias=dias, machine_id=machine_id)


@app.get("/carteiras", response_class=HTMLResponse)
def pagina_carteiras(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request=request, name="carteiras.html", context={"pagina_ativa": "carteiras"}
    )


@app.get("/dashboard", response_class=HTMLResponse)
def pagina_dashboard(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request=request, name="dashboard.html", context={"pagina_ativa": "dashboard"}
    )


@app.get("/api/cert-installer/diagnostico", dependencies=[Depends(require_admin)])
def diagnostico_do_instalador() -> dict:
    """
    Estado do módulo instalador, num lugar só.

    Cada bloco corresponde a algo que já falhou em produção sem aviso:

    - **binário**: `is_file()` só era consultado no clique, e a falha virava um
      503 com instrução de rebuild — inútil na Vercel, onde o FS é read-only.
    - **assinatura**: adiada em 11/08 e "não verificada em máquina real". Fica
      visível aqui em vez de dormir num changelog.
    - **cofre**: as seis falhas de instalação registradas têm causa única
      ("Senha ausente no cofre"), e descobrir isso exigiu ler o agent.log de
      uma máquina remota.
    - **chaves**: em 15/08 a chave foi trocada sem rotação e todo o cofre virou
      lixo cifrado. Nada na interface disse isso.
    """
    from app import pe_assinatura

    binario = pe_assinatura.inspecionar(INSTALADOR_AVULSO_EXE)
    out: dict = {
        "binario": {
            "existe": binario.existe,
            "caminho": binario.caminho,
            "tamanho_bytes": binario.tamanho_bytes,
            "sha256": binario.sha256,
            "modificado_em": binario.modificado_em,
            "assinado": binario.assinado,
            "assinatura_detalhe": binario.assinatura_detalhe,
            "signatarios": [vars(s) for s in binario.signatarios],
        },
        # O ícone é entrada de BUILD, não configuração de runtime: o binário é
        # o mesmo para todo download de propósito, e é isso que permite assiná-lo
        # uma vez e acumular reputação no SmartScreen. Trocar exige recompilar.
        "icone": {
            "arquivo": str(ROOT / "ico" / "icone.ico"),
            "presente": (ROOT / "ico" / "icone.ico").is_file(),
            "observacao": "Definido em Instalar_Certificado.spec; trocar exige recompilar e reassinar.",
        },
    }

    try:
        out["cofre"] = cert_installer.diagnostico_do_cofre()
        out["chaves"] = cert_installer.diagnostico_das_chaves()
    except Exception as e:  # noqa: BLE001
        logger.exception("Falha no diagnóstico do cofre")
        out["cofre"] = {"erro": str(e)}
        out["chaves"] = {"erro": str(e)}

    return out


@app.post("/api/cert-installer/revalidar-cofre", dependencies=[Depends(require_admin)])
def revalidar_cofre() -> dict:
    """
    Prova que a chave em vigor decifra o que está guardado.

    Contar linhas não prova nada: em 15/08 o cofre tinha uma linha íntegra, com
    todos os campos preenchidos, e completamente indecifrável. É `POST` porque
    faz trabalho criptográfico de verdade, não porque grave algo.
    """
    try:
        return {"resultados": cert_installer.revalidar_cofre()}
    except RuntimeError as e:
        raise HTTPException(status_code=503, detail=str(e))
    except Exception:
        logger.exception("Falha ao revalidar o cofre")
        raise HTTPException(status_code=500, detail="Erro interno ao revalidar o cofre")


@app.get("/api/cert-installer/instalabilidade")
def instalabilidade(
    machine_id: str = Query(..., min_length=1),
    token: auth.TokenData = Depends(require_auth),
) -> dict:
    """
    O que **este** usuário pode instalar nesta máquina, e o motivo de cada não.

    Alimenta a seleção do Início. Não é barreira — a barreira é
    `assegurar_carteira`, no momento de emitir o token. Isto existe para a tela
    não convidar o usuário a marcar o que o servidor vai recusar depois, com o
    erro chegando só na máquina dele.
    """
    user_id = _user_id_da_sessao(token)
    if not user_id:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    try:
        itens = cert_installer.estado_de_instalabilidade(machine_id, user_id, token.role)
    except (cert_installer.CustodiaIndisponivel, cert_installer.CarteiraIndisponivel) as e:
        logger.warning("Instalabilidade indisponível (%s): %s", machine_id, e)
        raise HTTPException(
            status_code=503,
            detail="Não foi possível verificar quais certificados estão disponíveis.",
        )
    return {
        "machine_id": machine_id,
        "alcance_total": (token.role or "").strip().lower()
        in cert_installer.PAPEIS_COM_ALCANCE_TOTAL,
        "itens": itens,
    }


# A rota POST /api/cert-installer/prepare foi removida em 16/08/2026.
#
# Ela instalava numa máquina que já roda o agente: enfileirava um comando, o
# agente pegava no poll seguinte e instalava lá. O cliente não usa esse caminho
# -- o operador baixa o .exe e instala na própria máquina --, e endpoint que
# emite token de instalação sem ninguém usar é superfície de ataque sem
# contrapartida: um token É a entrega da chave privada.
#
# `cert_installer.enqueue_install_command` continua existindo e o agente
# continua entendendo o comando `instalar_certificados`. Mexer no agente exige
# recompilar e reinstalar o .exe no ANALISESRV -- custo operacional real por
# zero benefício, já que sem quem enfileire o comando ele nunca chega.
#
# Se o caminho voltar a fazer sentido, o que falta é a rota: a barreira de
# carteira e a trilha já existem, e `tests/test_carteira.py` obriga qualquer
# rota nova que emita token a passar por elas.


class RedeemRequest(BaseModel):
    """Payload enviado pelo agente para resgatar o bundle criptografado."""
    token: str
    clientPublicKey: str  # SPKI base64 (ECDH P-256)


@app.post("/api/cert-installer/redeem")
def redeem_install(
    body: RedeemRequest,
    request: Request,
    _token: auth.TokenData = Depends(require_agent_or_admin),
):
    """
    Agente consome o token e recebe o bundle de certificados criptografado via ECDH.
    """
    client_ip = request.client.host if request.client else None

    # 1. Validar e consumir token
    token_data = cert_installer.validate_and_consume_token(body.token)
    if not token_data:
        raise HTTPException(status_code=403, detail="Token inválido, expirado ou já consumido")

    user_id = token_data["user_id"]
    token_id = str(token_data["id"])
    cert_ids = token_data.get("certificate_ids") or []

    # 2. Log REDIMIDO
    cert_installer.log_event(
        event="REDIMIDO",
        user_id=user_id,
        user_email=token_data.get("user_email"),
        token_id=token_id,
        target_machine=token_data.get("target_machine"),
        client_ip=client_ip,
    )

    # 3. Montar bundle criptografado
    try:
        bundle = cert_installer.build_encrypted_bundle(
            certificate_ids=cert_ids,
            client_public_key_b64=body.clientPublicKey,
        )
        return bundle
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception:
        logger.exception("Erro ao montar bundle criptografado")
        raise HTTPException(status_code=500, detail="Erro interno ao montar bundle")


# ── Instalador avulso (modelo Ninite) ─────────────────────────────────────
#
# O agente resgata em /redeem autenticando-se com X-API-Key. O instalador que o
# usuário baixa não pode fazer o mesmo: embutir a API key no executável seria
# distribuí-la a todos que baixassem — e ela abre todas as rotas do agente.
#
# Aqui o próprio token É a credencial, como numa URL de download assinada. Isso
# se sustenta porque ele é de uso único (compare-and-swap em
# validate_and_consume_token), expira em CERT_INSTALL_TOKEN_TTL_MIN minutos, e
# só libera os certificate_ids gravados nele. O que falta a um bearer assim é
# resistência a força bruta, daí o limite por IP abaixo.

_CLAIM_JANELA_SEC = 60
_CLAIM_MAX_POR_JANELA = 10
_claim_tentativas: dict[str, list[float]] = {}
_claim_lock = threading.Lock()


def _claim_rate_limit(ip: str) -> bool:
    """True se o IP ainda pode tentar. Janela deslizante em memória."""
    agora = time.time()
    with _claim_lock:
        tentativas = [t for t in _claim_tentativas.get(ip, []) if agora - t < _CLAIM_JANELA_SEC]
        if len(tentativas) >= _CLAIM_MAX_POR_JANELA:
            _claim_tentativas[ip] = tentativas
            return False
        tentativas.append(agora)
        _claim_tentativas[ip] = tentativas
        # A limpeza oportunista evita o dicionário crescer sem limite com IPs
        # que apareceram uma vez. Barato: só roda quando o mapa já está grande.
        if len(_claim_tentativas) > 1000:
            for k in [k for k, v in _claim_tentativas.items()
                      if not v or agora - v[-1] > _CLAIM_JANELA_SEC]:
                _claim_tentativas.pop(k, None)
        return True


@app.post("/api/cert-installer/claim")
def claim_install(body: RedeemRequest, request: Request):
    """
    Resgate SEM API key, para o instalador avulso. O token é a credencial.

    Deliberadamente não distingue token inválido de expirado ou já consumido:
    a resposta única evita virar oráculo de tokens válidos.
    """
    client_ip = request.client.host if request.client else "desconhecido"

    if not _claim_rate_limit(client_ip):
        raise HTTPException(status_code=429, detail="Muitas tentativas. Aguarde um minuto.")

    token_data = cert_installer.validate_and_consume_token(body.token)
    if not token_data:
        raise HTTPException(status_code=403, detail="Token inválido, expirado ou já utilizado")

    cert_installer.log_event(
        event="REDIMIDO",
        user_id=token_data["user_id"],
        user_email=token_data.get("user_email"),
        token_id=str(token_data["id"]),
        target_machine=token_data.get("target_machine"),
        client_ip=client_ip,
        detail="instalador avulso",
    )

    try:
        return cert_installer.build_encrypted_bundle(
            certificate_ids=token_data.get("certificate_ids") or [],
            client_public_key_b64=body.clientPublicKey,
        )
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception:
        logger.exception("Erro ao montar bundle para instalador avulso")
        raise HTTPException(status_code=500, detail="Erro interno ao montar bundle")


class PrepararDownloadRequest(BaseModel):
    """Admin escolhe os certificados e recebe o link do instalador."""
    certificate_ids: List[str]
    nome: Optional[str] = None


@app.post("/api/cert-installer/preparar-download")
def preparar_download(
    body: PrepararDownloadRequest,
    request: Request,
    token: auth.TokenData = Depends(require_auth),
):
    """
    Cria o token e devolve a URL do instalador — sem enfileirar nada.

    Difere de /prepare no destino: lá o alvo é uma máquina que já roda o agente,
    e o comando vai para a fila dela. Aqui o alvo é a máquina de quem clicou,
    que não tem agente nenhum; o "transporte" é o próprio download.

    Mesma barreira de carteira da outra: são os dois caminhos que produzem um
    token de instalação, e um token é o que entrega a chave privada.
    """
    if not body.certificate_ids:
        raise HTTPException(status_code=400, detail="Selecione ao menos um certificado")

    user_id = _user_id_da_sessao(token)
    if not user_id:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")

    _validar_pedido_de_instalacao(user_id, token.role, body.certificate_ids)

    client_ip = request.client.host if request.client else None

    try:
        token_raw, token_id, expires_at = cert_installer.create_install_token(
            user_id=user_id,
            user_email=token.email,
            # Não há máquina alvo conhecida: quem executar o instalador define
            # onde o certificado entra. Fica registrado como tal na auditoria.
            target_machine="download-avulso",
            certificate_ids=body.certificate_ids,
            client_ip=client_ip,
        )
        for cid in body.certificate_ids:
            cert_installer.log_event(
                event="SOLICITADO",
                user_id=user_id,
                user_email=token.email,
                token_id=token_id,
                certificate_id=cid,
                target_machine="download-avulso",
                client_ip=client_ip,
            )
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception:
        logger.exception("Erro ao preparar download do instalador")
        raise HTTPException(status_code=500, detail="Erro interno ao preparar instalador")

    nome = quote((body.nome or "Certificado")[:60])
    return {
        "status": "ok",
        "download_url": f"/instalador/baixar/{token_raw}?nome={nome}",
        "expires_at": expires_at.isoformat(),
        "validade_min": config.CERT_INSTALL_TOKEN_TTL_MIN,
    }


# Onde o build deposita o instalador avulso já compilado.
INSTALADOR_AVULSO_EXE = ROOT / "dist" / "Instalar_Certificado.exe"

# Só o alfabeto de secrets.token_urlsafe. Serve para recusar, antes de tocar o
# disco, qualquer coisa que não tenha forma de token — inclusive travessia de
# caminho, já que o valor vai para o nome do arquivo servido.
_TOKEN_SEGURO = re.compile(r"^[A-Za-z0-9_-]{16,128}$")


@app.get("/instalador/baixar/{token}")
def baixar_instalador(token: str, nome: str = Query("Certificado")):
    """
    Serve o instalador avulso com o token no NOME do arquivo.

    O binário é sempre o mesmo — não é recompilado por download. Isso é o que
    permite assiná-lo uma vez e acumular reputação no SmartScreen; um executável
    gerado a cada clique seria inédito para o Windows em toda instalação, e o
    aviso de "aplicativo não reconhecido" apareceria justamente no fluxo que
    deveria ser de um clique só. O instalador lê o token do próprio argv[0].
    """
    if not _TOKEN_SEGURO.match(token):
        raise HTTPException(status_code=400, detail="Token malformado")

    if not INSTALADOR_AVULSO_EXE.is_file():
        raise HTTPException(
            status_code=503,
            detail="Instalador ainda não compilado no servidor. Rode scripts/build_instalador_avulso.ps1.",
        )

    # O template é configurável desde 15/08, mas o `{token}` continua sendo
    # obrigatório: o instalador o lê do próprio argv[0]. `montar_nome_do_arquivo`
    # cai no padrão se o template gravado for inválido — entregar um instalador
    # que funciona com nome padrão é melhor que um nome bonito que não instala.
    try:
        template = load_settings().instalador_nome_template
    except Exception:  # noqa: BLE001
        logger.exception("Falha ao ler o template do nome; usando o padrão")
        template = ""

    return FileResponse(
        path=INSTALADOR_AVULSO_EXE,
        media_type="application/vnd.microsoft.portable-executable",
        filename=cert_installer.montar_nome_do_arquivo(template, nome, token),
    )


class InstallResultItem(BaseModel):
    certificateId: str
    fingerprint: Optional[str] = None
    thumbprint: Optional[str] = None
    status: str  # "OK" | "FALHA"
    detail: Optional[str] = None


class ReportRequest(BaseModel):
    """Payload enviado pelo agente após instalar os certificados."""
    token: str
    results: List[InstallResultItem]


@app.post("/api/cert-installer/report")
def report_install(
    body: ReportRequest,
    request: Request,
    _token: auth.TokenData = Depends(require_agent_or_admin),
):
    """
    Agente reporta o resultado da instalação de cada certificado.
    """
    return _registrar_relatorio(body, request)


@app.post("/api/cert-installer/report-avulso")
def report_install_avulso(body: ReportRequest, request: Request):
    """
    Mesma gravação, para o instalador avulso — que não tem API key.

    Abrir a rota não afeta o gate real: `_registrar_relatorio` já exigia um
    token conhecido E já redimido, o que a dependência de autenticação apenas
    duplicava. O pior que um token vazado permite aqui é sujar a auditoria da
    própria instalação que ele representa; não devolve material criptográfico.
    """
    ip = request.client.host if request.client else "desconhecido"
    if not _claim_rate_limit(ip):
        raise HTTPException(status_code=429, detail="Muitas tentativas. Aguarde um minuto.")
    return _registrar_relatorio(body, request)


def _registrar_relatorio(body: ReportRequest, request: Request) -> dict:
    client_ip = request.client.host if request.client else None

    # Validar que o token existe e foi redimido (consumed_at != null)
    import hashlib
    token_hash = hashlib.sha256(body.token.encode()).hexdigest()
    from app.settings_state import _supabase
    sb = _supabase()
    if not sb:
        raise HTTPException(status_code=500, detail="Supabase não configurado")

    try:
        r = sb.table("install_token").select("*").eq("token_hash", token_hash).execute()
        rows = r.data or []
        if not rows:
            raise HTTPException(status_code=403, detail="Token desconhecido")
        token_row = rows[0]
        if not token_row.get("consumed_at"):
            raise HTTPException(status_code=403, detail="Token ainda não foi redimido")
    except HTTPException:
        raise
    except Exception:
        logger.exception("Erro ao verificar token no report")
        raise HTTPException(status_code=500, detail="Erro interno")

    user_id = token_row["user_id"]
    token_id = str(token_row["id"])

    # Gravar log para cada resultado
    for result in body.results:
        event = "CONCLUIDO" if result.status.upper() == "OK" else "ERRO"
        cert_installer.log_event(
            event=event,
            user_id=user_id,
            user_email=token_row.get("user_email"),
            token_id=token_id,
            certificate_id=result.certificateId,
            fingerprint=result.fingerprint,
            target_machine=token_row.get("target_machine"),
            status=result.status,
            detail=result.detail,
            client_ip=client_ip,
        )

    return {
        "status": "ok",
        "processed": len(body.results),
    }


# ── Endpoints auxiliares do instalador ────────────────────────────────────

@app.get("/api/cert-installer/available")
def list_available_certificates(
    machine_id: Optional[str] = Query(None),
    token: auth.TokenData = Depends(require_admin),
):
    """Lista certificados PFX disponíveis para instalação (sem dados cifrados)."""
    certs = cert_installer.list_available_pfx(machine_id=machine_id)
    return {
        "certificates": [
            {
                "id": c.id,
                "fingerprint": c.fingerprint,
                "machine_id": c.machine_id,
                "nome_titular": c.nome_titular,
                "documento": c.documento,
                "documento_tipo": c.documento_tipo,
                "subject": c.subject,
                "not_before": c.not_before,
                "not_after": c.not_after,
                "friendly_name": c.friendly_name,
                "uploaded_at": c.uploaded_at,
            }
            for c in certs
        ]
    }


@app.get("/api/cert-installer/logs")
def list_installer_logs(
    limit: int = Query(100, ge=1, le=500),
    token: auth.TokenData = Depends(require_admin),
):
    """Lista logs de auditoria de instalação."""
    logs = cert_installer.list_install_logs(limit=limit)
    return {"logs": logs}


@app.get("/api/cert-installer/trilha", dependencies=[Depends(require_admin)])
def trilha_de_instalacao(
    dias: int = Query(30, ge=1, le=365),
    user_email: Optional[str] = Query(None),
    apenas_falhas: bool = Query(False),
    limite: int = Query(500, ge=1, le=1000),
) -> dict:
    """
    A trilha agrupada por token: uma linha por tentativa de instalação.

    Substitui a lista plana, que mostrava eventos soltos em ordem cronológica.
    O que importa é **onde a cadeia quebrou** — e os números de produção
    mostram por quê: nove tentativas, seis mortas no mesmo ponto e pela mesma
    causa. Em fila, são 25 linhas sem forma.
    """
    desde = (datetime.now(timezone.utc) - timedelta(days=dias)).isoformat()
    cadeias = cert_installer.cadeias_de_instalacao(
        limite=limite, desde=desde, user_email=user_email, apenas_com_falha=apenas_falhas
    )
    return {
        "dias": dias,
        "resumo": cert_installer.resumo_das_cadeias(cadeias),
        "cadeias": cadeias,
    }


@app.post("/api/cert-installer/cleanup")
def cleanup_tokens(token: auth.TokenData = Depends(require_admin)):
    """Remove tokens de instalação expirados (manutenção)."""
    count = cert_installer.cleanup_expired_tokens()
    return {"status": "ok", "removed": count}


def _resolve_user_id(email: str) -> Optional[str]:
    """Busca o UUID do usuário pelo email."""
    from app.settings_state import _supabase
    sb = _supabase()
    if not sb:
        return None
    try:
        r = sb.table("users").select("id").eq("email", email).limit(1).execute()
        if r.data:
            return str(r.data[0]["id"])
    except Exception:
        logger.exception("Erro ao resolver user_id para email=%s", email)
    return None


# ── Página HTML do Instalador ─────────────────────────────────────────────

@app.get("/instalador", response_class=HTMLResponse)
def page_instalador(request: Request) -> HTMLResponse:
    # Assinatura por keyword, igual às outras 8 rotas de página. A forma
    # posicional antiga — TemplateResponse(name, context) — quebra nesta versão
    # do Starlette com "TypeError: unhashable type: 'dict'", e a página do
    # módulo respondia 500 em toda requisição.
    # O nonce vem de request.state.nonce no template, como nos demais.
    return templates.TemplateResponse(
        request=request, name="instalador.html", context={"pagina_ativa": "instalador"}
    )

